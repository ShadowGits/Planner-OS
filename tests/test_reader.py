from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import TestCase

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from planner_engine import (
    ExcelPlannerStore,
    PlannerEngine,
    PlannerWorkbookError,
    Priority,
    TaskStatus,
)


def create_representative_planner(path: Path) -> None:
    """Create a small workbook with the planner's repeated month structure."""

    workbook = Workbook()
    july = workbook.active
    july.title = "Jul 2026"
    july["B3"] = "MONTHLY GOALS · WEEKLY BREAKDOWN · CUSTOM ITEMS"
    july["D6"] = '=COUNTIF(K10:K200,"Done")'
    july["B8"] = "▸  MONTHLY GOALS"
    july["A9"] = "#"
    july["B9"] = "GOAL / TASK"
    july["C9"] = "CATEGORY"
    july["D9"] = "PRIORITY"
    july["E9"] = "TARGET WEEK"
    july["H9"] = "STATUS"
    july["L9"] = "NOTES"
    july.merge_cells("E9:G9")
    july.merge_cells("H9:K9")
    july["B10"] = "Finish IELTS essay"
    july["C10"] = "IELTS"
    july["D10"] = "High"
    july["E10"] = "Week 1"
    july["H10"] = "In Progress"
    july["L10"] = "Draft intro"
    july.merge_cells("E10:G10")
    july.merge_cells("H10:K10")
    july["B11"] = None
    july["C11"] = "Template category"
    july.merge_cells("E11:G11")
    july.merge_cells("H11:K11")

    july["B20"] = "WEEK 1  ·  29 Jun – 05 Jul 2026"
    july["A21"] = "#"
    july["B21"] = "TASK / GOAL"
    july["C21"] = "CATEGORY"
    july["D21"] = "Mon\n29 Jun"
    july["K21"] = "STATUS"
    july["L21"] = "NOTES"
    july["B22"] = "  ↳ From monthly goals (fill in tasks planned for this week)"
    july["B23"] = "Book German lesson"
    july["C23"] = "German"
    july["K23"] = "Done"
    july["L23"] = "Booked with tutor"
    july["C24"] = "Blank template row should be ignored"

    july["B35"] = "WEEK 2  ·  06 Jul – 12 Jul 2026"
    july["A36"] = "#"
    july["B36"] = "TASK / GOAL"
    july["C36"] = "CATEGORY"
    july["K36"] = "STATUS"
    july["L36"] = "NOTES"
    july["B37"] = "  ↳ From monthly goals (fill in tasks planned for this week)"
    july["B38"] = "Gym double class"
    july["C38"] = "Gym"
    july["K38"] = "Not Started"
    july["L38"] = "Dance + strength"

    status_validation = DataValidation(
        type="list",
        formula1='"Done,In Progress,To Do,Blocked,Skipped"',
    )
    status_validation.add("H10:H17")
    status_validation.add("K23:K28")
    status_validation.add("K38:K43")
    july.add_data_validation(status_validation)

    priority_validation = DataValidation(type="list", formula1='"High,Medium,Low"')
    priority_validation.add("D10:D17")
    july.add_data_validation(priority_validation)

    category_validation = DataValidation(
        type="list",
        formula1='"Work,Learning,Health,Personal,Finance,German,Reading,Other"',
    )
    category_validation.add("C10:C17")
    category_validation.add("C23:C28")
    category_validation.add("C38:C43")
    july.add_data_validation(category_validation)

    target_week_validation = DataValidation(
        type="list",
        formula1='"Week 1,Week 2,Week 3,Week 4,Week 5,All Month"',
    )
    target_week_validation.add("E10:E17")
    july.add_data_validation(target_week_validation)

    august = workbook.create_sheet("Aug 2026")
    august["B8"] = "▸  MONTHLY GOALS"
    august["B9"] = "GOAL / TASK"
    august["C9"] = "CATEGORY"
    august["D9"] = "PRIORITY"
    august["E9"] = "TARGET WEEK"
    august["H9"] = "STATUS"
    august["L9"] = "NOTES"
    august.merge_cells("E9:G9")
    august.merge_cells("H9:K9")
    august["B20"] = "WEEK 1  ·  27 Jul – 02 Aug 2026"
    august["B21"] = "TASK / GOAL"
    august["C21"] = "CATEGORY"
    august["K21"] = "STATUS"
    august["L21"] = "NOTES"

    workbook.create_sheet("Settings")
    workbook.save(path)
    workbook.close()


class PlannerReaderTests(TestCase):
    """Tests for read-only planner concept parsing."""

    def test_monthly_sheets_are_listed(self) -> None:
        with TemporaryDirectory() as directory:
            planner_path = Path(directory) / "planner.xlsx"
            create_representative_planner(planner_path)
            engine = PlannerEngine(ExcelPlannerStore(planner_path, Path(directory)))

            self.assertEqual(engine.list_months(), ["Jul 2026", "Aug 2026"])

    def test_monthly_goals_are_parsed(self) -> None:
        with TemporaryDirectory() as directory:
            planner_path = Path(directory) / "planner.xlsx"
            create_representative_planner(planner_path)
            engine = PlannerEngine(ExcelPlannerStore(planner_path, Path(directory)))

            goals = engine.get_monthly_goals("Jul 2026")

            self.assertEqual(len(goals), 1)
            self.assertEqual(goals[0].name, "Finish IELTS essay")
            self.assertEqual(goals[0].category, "IELTS")
            self.assertEqual(goals[0].priority, Priority.HIGH)
            self.assertEqual(goals[0].target_week, "Week 1")
            self.assertEqual(goals[0].status, TaskStatus.IN_PROGRESS)
            self.assertEqual(goals[0].notes, "Draft intro")
            self.assertEqual(goals[0].sheet_name, "Jul 2026")
            self.assertEqual(goals[0].row_number, 10)
            self.assertEqual(goals[0].cell_references["name"], "B10")
            self.assertEqual(goals[0].cell_references["status"], "H10")
            self.assertEqual(goals[0].cell_references["target_week"], "E10")

    def test_weekly_sections_and_tasks_are_parsed(self) -> None:
        with TemporaryDirectory() as directory:
            planner_path = Path(directory) / "planner.xlsx"
            create_representative_planner(planner_path)
            engine = PlannerEngine(ExcelPlannerStore(planner_path, Path(directory)))

            sections = engine.get_week_sections("Jul 2026")

            self.assertEqual(len(sections), 2)
            self.assertEqual(sections[0].name, "WEEK 1")
            self.assertEqual(sections[0].heading_row, 20)
            self.assertEqual(sections[0].header_row, 21)
            self.assertEqual(len(sections[0].tasks), 1)

            task = sections[0].tasks[0]
            self.assertEqual(task.name, "Book German lesson")
            self.assertEqual(task.category, "German")
            self.assertEqual(task.priority, Priority.UNKNOWN)
            self.assertEqual(task.status, TaskStatus.DONE)
            self.assertEqual(task.notes, "Booked with tutor")
            self.assertEqual(task.row_number, 23)
            self.assertEqual(task.sheet_name, "Jul 2026")
            self.assertEqual(task.week_name, "WEEK 1")
            self.assertEqual(task.cell_references["name"], "B23")
            self.assertEqual(task.cell_references["category"], "C23")
            self.assertEqual(task.cell_references["status"], "K23")
            self.assertEqual(task.cell_references["notes"], "L23")

    def test_blank_template_rows_are_ignored(self) -> None:
        with TemporaryDirectory() as directory:
            planner_path = Path(directory) / "planner.xlsx"
            create_representative_planner(planner_path)
            engine = PlannerEngine(ExcelPlannerStore(planner_path, Path(directory)))

            month_plan = engine.get_month_plan("Jul 2026")
            task_names = [
                task.name
                for section in month_plan.week_sections
                for task in section.tasks
            ]

            self.assertEqual(
                task_names,
                ["Book German lesson", "Gym double class"],
            )

    def test_task_lookup_is_case_insensitive(self) -> None:
        with TemporaryDirectory() as directory:
            planner_path = Path(directory) / "planner.xlsx"
            create_representative_planner(planner_path)
            engine = PlannerEngine(ExcelPlannerStore(planner_path, Path(directory)))

            task = engine.find_task("jul 2026", "book german LESSON")

            self.assertIsNotNone(task)
            self.assertEqual(task.name, "Book German lesson")

    def test_unknown_month_raises_clear_error(self) -> None:
        with TemporaryDirectory() as directory:
            planner_path = Path(directory) / "planner.xlsx"
            create_representative_planner(planner_path)
            engine = PlannerEngine(ExcelPlannerStore(planner_path, Path(directory)))

            with self.assertRaisesRegex(
                PlannerWorkbookError,
                "Unknown planner month: NotAMonth",
            ):
                engine.get_month_plan("NotAMonth")
