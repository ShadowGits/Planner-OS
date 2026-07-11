from __future__ import annotations

import json
import os
import subprocess
import sys
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import TestCase

from openpyxl import load_workbook

from test_writer import create_writer_workbook


class ShadowCLITests(TestCase):
    """Tests for the Shadow command-line interface."""

    def test_validate(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            create_writer_workbook(planner_path)

            result = self._run_shadow(tmp_path, planner_path, "validate")

            self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
            self.assertIn("Everything OK", result.stdout)

    def test_import(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            input_path = tmp_path / "import.json"
            create_writer_workbook(planner_path)
            input_path.write_text(
                json.dumps(
                    {
                        "month": "Jul 2026",
                        "monthly_goals": [
                            {
                                "goal": "Imported Goal",
                                "category": "Learning",
                                "priority": "High",
                                "target_week": "W2",
                                "status": "Not Started",
                                "notes": "",
                            }
                        ],
                        "weeks": [
                            {
                                "week": 2,
                                "tasks": [
                                    {
                                        "task": "Imported Task",
                                        "category": "Study",
                                        "status": "Not Started",
                                        "notes": "",
                                    }
                                ],
                            }
                        ],
                    }
                ),
                encoding="utf-8",
            )

            result = self._run_shadow(tmp_path, planner_path, "import", str(input_path))

            self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
            self.assertIn("Import complete", result.stdout)
            workbook = load_workbook(planner_path)
            try:
                self.assertEqual(workbook["Jul 2026"]["B11"].value, "Imported Goal")
                self.assertEqual(workbook["Jul 2026"]["B39"].value, "Imported Task")
            finally:
                workbook.close()

    def test_plan_today(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            create_writer_workbook(planner_path)

            result = self._run_shadow(tmp_path, planner_path, "plan", "today")

            self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
            self.assertIn("Shadow plan", result.stdout)

    def test_complete(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            create_writer_workbook(planner_path)

            result = self._run_shadow(
                tmp_path,
                planner_path,
                "complete",
                "Existing Task",
            )

            self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
            self.assertIn("Task complete", result.stdout)
            self.assertIn("Progress updated: yes", result.stdout)
            workbook = load_workbook(planner_path)
            try:
                self.assertEqual(workbook["Jul 2026"]["K38"].value, "Done")
            finally:
                workbook.close()

    def test_status(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            create_writer_workbook(planner_path)

            result = self._run_shadow(tmp_path, planner_path, "status")

            self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
            self.assertIn("Shadow status", result.stdout)
            self.assertIn("Gym:", result.stdout)
            self.assertIn("Active slippage alerts:", result.stdout)

    def test_rules_list_and_set_work_days(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            rules_path = tmp_path / "rules.yaml"
            create_writer_workbook(planner_path)
            rules_path.write_text(
                (Path(__file__).resolve().parents[1] / "config" / "rules.yaml").read_text(
                    encoding="utf-8"
                ),
                encoding="utf-8",
            )

            listed = self._run_shadow(
                tmp_path,
                planner_path,
                "rules",
                "list",
                rules_path=rules_path,
            )
            updated = self._run_shadow(
                tmp_path,
                planner_path,
                "rules",
                "set-work-days",
                "Monday",
                "Tuesday",
                rules_path=rules_path,
            )

            self.assertEqual(listed.returncode, 0, listed.stdout + listed.stderr)
            self.assertIn("active_days:", listed.stdout)
            self.assertEqual(updated.returncode, 0, updated.stdout + updated.stderr)
            self.assertIn("Work days updated", updated.stdout)
            persisted = rules_path.read_text(encoding="utf-8")
            self.assertIn("- Monday", persisted)
            self.assertIn("- Sunday", persisted)

    def _run_shadow(
        self,
        tmp_path: Path,
        planner_path: Path,
        *args: str,
        rules_path: Path | None = None,
    ) -> subprocess.CompletedProcess[str]:
        repo_root = Path(__file__).resolve().parents[1]
        python_bin = Path(sys.executable).parent
        env = {
            **os.environ,
            "PATH": f"{repo_root}{os.pathsep}{python_bin}{os.pathsep}{os.environ['PATH']}",
        }
        return subprocess.run(
            [
                "shadow",
                "--planner",
                str(planner_path),
                "--backup-dir",
                str(tmp_path / "backups"),
                "--rules",
                str(rules_path or (repo_root / "config" / "rules.yaml")),
                "--month",
                "Jul 2026",
                *args,
            ],
            cwd=tmp_path,
            env=env,
            check=False,
            capture_output=True,
            text=True,
        )
