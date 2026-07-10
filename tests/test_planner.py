import subprocess
import sys
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import TestCase

from openpyxl import Workbook, load_workbook

from planner_engine import ExcelPlannerStore, PlannerEngine


def create_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Plan"
    worksheet["A1"] = "Original"
    workbook.save(path)
    workbook.close()


class PlannerEngineTests(TestCase):
    """Tests for the high-level Planner Engine API."""

    def test_planner_engine_backs_up_before_write(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            backup_dir = tmp_path / "backups"
            create_workbook(planner_path)
            engine = PlannerEngine(ExcelPlannerStore(planner_path, backup_dir))

            result = engine.write_cell("Plan", "A1", "Updated")

            self.assertTrue(result.backup_path.exists())
            self.assertEqual(engine.read_cell("Plan", "A1"), "Updated")
            backup_workbook = load_workbook(result.backup_path)
            try:
                self.assertEqual(backup_workbook["Plan"]["A1"].value, "Original")
            finally:
                backup_workbook.close()

    def test_edit_planner_cli_uses_planner_engine(self) -> None:
        with TemporaryDirectory() as directory:
            repo_root = Path(__file__).resolve().parents[1]
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            backup_dir = tmp_path / "backups"
            create_workbook(planner_path)

            result = subprocess.run(
                [
                    sys.executable,
                    str(repo_root / "src" / "edit_planner.py"),
                    "--sheet",
                    "Plan",
                    "--cell",
                    "A1",
                    "--value",
                    "CLI Updated",
                    "--planner-path",
                    str(planner_path),
                    "--backup-dir",
                    str(backup_dir),
                ],
                cwd=repo_root,
                check=True,
                capture_output=True,
                text=True,
            )

            workbook = load_workbook(planner_path)
            try:
                self.assertEqual(workbook["Plan"]["A1"].value, "CLI Updated")
            finally:
                workbook.close()

            self.assertIn("Backup:", result.stdout)
            self.assertIn("Updated Plan!A1 = CLI Updated", result.stdout)
            self.assertEqual(len(list(backup_dir.glob("*.xlsx"))), 1)
