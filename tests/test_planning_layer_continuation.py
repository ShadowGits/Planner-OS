from __future__ import annotations

from datetime import date, datetime, timedelta
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import TestCase
from unittest.mock import Mock

from openpyxl import load_workbook

from planner_engine import ExcelPlannerStore, PlannerEngine, ProgressEngine, RulesEngine, Writer
from planner_engine.calendar_sync import CalendarSyncService
from planner_engine.checkin import DailyCheckInService
from planner_engine.command_router import PlannerCommand, PlannerCommandRouter, parse_common_intent
from planner_engine.models import DailyPlan, MonthPlan, ScheduledBlock
from planner_engine.monthly_planner import MonthlyPlanPreview
from planner_engine.planning_commands import PlanningCommandService
from planner_integrations.google_calendar import CalendarSyncResult, GoogleCalendarClient
from test_google_calendar import FakeService
from test_writer import create_writer_workbook, writer_for


TARGET = date(2026, 7, 11)


def dated_workbook(path: Path) -> None:
    create_writer_workbook(path)
    workbook = load_workbook(path)
    try:
        sheet = workbook["Jul 2026"]
        sheet["B50"] = "WEEK 3  ·  13 Jul – 19 Jul 2026"
        sheet["A51"] = "#"
        sheet["B51"] = "TASK / GOAL"
        sheet["C51"] = "CATEGORY"
        sheet["K51"] = "STATUS"
        sheet["L51"] = "NOTES"
        sheet["B52"] = "  ↳ From monthly goals (fill in tasks planned for this week)"
        for header_row, start in (
            (21, date(2026, 6, 29)),
            (36, date(2026, 7, 6)),
            (51, date(2026, 7, 13)),
        ):
            for offset in range(7):
                sheet.cell(header_row, 4 + offset).value = (start + timedelta(days=offset)).strftime("%d %b")
        workbook.save(path)
    finally:
        workbook.close()


def services(tmp_path: Path):
    planner_path = tmp_path / "planner.xlsx"
    dated_workbook(planner_path)
    writer, progress = writer_for(planner_path, tmp_path / "backups")
    engine = writer.engine
    rules = RulesEngine(Path(__file__).resolve().parents[1] / "config" / "rules.yaml")
    return planner_path, engine, writer, progress, rules


class RecordingCalendar:
    def __init__(self) -> None:
        self.calls = []

    def sync_plan(self, plan, *, start, end, scope):
        self.calls.append((plan, start, end, scope))
        return CalendarSyncResult(True, 1, 0, 0, 0, [], [])


class EmptyScheduler:
    def plan_day(self, month_plan, target_date):
        del month_plan
        return DailyPlan(target_date, [], [])


class DatedTaskTests(TestCase):
    def test_add_list_update_delete_exact_date(self) -> None:
        with TemporaryDirectory() as directory:
            _, _, writer, _, _ = services(Path(directory))
            result = writer.add_dated_task(TARGET, "Call plumber", 30, preferred_daypart="morning")
            self.assertTrue(result.success, result.errors)
            tasks = writer.list_dated_tasks(TARGET)
            self.assertEqual([(task.date, task.title) for task in tasks], [(TARGET, "Call plumber")])
            self.assertNotIn("Call plumber", [task.title for task in writer.list_dated_tasks(TARGET - timedelta(days=1))])

            updated = writer.update_dated_task(tasks[0].id, title="Call electrician", estimated_minutes=45)
            self.assertTrue(updated.success, updated.errors)
            tasks = writer.list_dated_tasks(TARGET)
            self.assertEqual((tasks[0].title, tasks[0].estimated_minutes), ("Call electrician", 45))

            deleted = writer.delete_dated_task(tasks[0].id)
            self.assertTrue(deleted.success, deleted.errors)
            self.assertEqual(writer.list_dated_tasks(TARGET), [])

    def test_batch_split_creates_two_exact_dates(self) -> None:
        with TemporaryDirectory() as directory:
            _, _, writer, _, _ = services(Path(directory))
            command = parse_common_intent(
                "sort my room half today and half tomorrow afternoon", today=date(2026, 7, 10)
            )
            result = writer.add_dated_tasks(command.payload["tasks"])
            self.assertTrue(result.success, result.errors)
            self.assertEqual(len(writer.list_dated_tasks(date(2026, 7, 10))), 1)
            self.assertEqual(len(writer.list_dated_tasks(date(2026, 7, 11))), 1)

    def test_batch_add_ids_are_distinct_and_addressable(self) -> None:
        with TemporaryDirectory() as directory:
            _, _, writer, _, _ = services(Path(directory))
            result = writer.add_dated_tasks([
                {"date": TARGET, "title": "Task A", "estimated_minutes": 30},
                {"date": TARGET, "title": "Task B", "estimated_minutes": 30},
                {"date": TARGET, "title": "Task C", "estimated_minutes": 30},
            ])
            self.assertTrue(result.success, result.errors)

            tasks = writer.list_dated_tasks(TARGET)
            ids = {task.id for task in tasks}
            # Every task in one batch must get its own findable id, or later
            # updates/deletes that address a task by id would break.
            self.assertEqual(len(ids), 3)
            self.assertEqual({task.title for task in tasks}, {"Task A", "Task B", "Task C"})

            by_title = {task.title: task for task in tasks}
            moved = writer.update_dated_task(by_title["Task B"].id, title="Task B2")
            self.assertTrue(moved.success, moved.errors)
            deleted = writer.delete_dated_task(by_title["Task C"].id)
            self.assertTrue(deleted.success, deleted.errors)

            remaining = {task.title for task in writer.list_dated_tasks(TARGET)}
            self.assertEqual(remaining, {"Task A", "Task B2"})

    def test_overwrite_preserves_manual_dated_task(self) -> None:
        with TemporaryDirectory() as directory:
            _, _, writer, _, _ = services(Path(directory))
            self.assertTrue(writer.add_dated_task(TARGET, "Manual task", 30).success)
            self.assertTrue(
                writer.add_dated_tasks(
                    [{"date": TARGET, "title": "Generated old", "estimated_minutes": 30,
                      "notes": "Planner OS generated plan"}]
                ).success
            )
            result = writer.add_dated_tasks(
                [{"date": TARGET, "title": "Generated new", "estimated_minutes": 30,
                  "notes": "Planner OS generated plan"}],
                overwrite_generated=True,
                overwrite_start=TARGET,
                overwrite_end=TARGET,
            )
            self.assertTrue(result.success, result.errors)
            titles = {task.title for task in writer.list_dated_tasks(TARGET)}
            self.assertEqual(titles, {"Manual task", "Generated new"})


class CalendarDateSyncTests(TestCase):
    def test_sync_date_is_exact_and_includes_dated_task(self) -> None:
        with TemporaryDirectory() as directory:
            _, engine, writer, _, rules = services(Path(directory))
            self.assertTrue(writer.add_dated_task(TARGET, "File ITR", 90, preferred_daypart="afternoon").success)
            client = RecordingCalendar()
            service = CalendarSyncService(engine, rules, client, EmptyScheduler())
            result = service.calendar_sync_date(TARGET)

            self.assertEqual((result.start_date, result.end_date), (TARGET.isoformat(), TARGET.isoformat()))
            self.assertIn(f"{TARGET.isoformat()} to {TARGET.isoformat()}", result.message)
            plan, start, end, scope = client.calls[0]
            self.assertEqual((start, end, scope), (TARGET, TARGET, "date"))
            self.assertEqual([day.date for day in plan.days], [TARGET])
            self.assertIn("File ITR", [block.title for block in plan.days[0].blocks])

    def test_moving_dated_task_preserves_id_and_updates_calendar_without_duplicate(self) -> None:
        with TemporaryDirectory() as directory:
            tmp = Path(directory)
            _, engine, writer, _, rules = services(tmp)
            old_date = date(2026, 7, 12)
            new_date = date(2026, 7, 13)
            add = writer.add_dated_task(old_date, "Call plumber", 30, preferred_daypart="morning")
            self.assertTrue(add.success, add.errors)
            original = writer.list_dated_tasks(old_date)[0]

            client = GoogleCalendarClient(
                service=FakeService(),
                decision_log=writer.decision_log,
                external_links_path=tmp / "external-links.json",
            )
            service = CalendarSyncService(engine, rules, client, EmptyScheduler())
            first_sync = service.calendar_sync_date(old_date)
            self.assertTrue(first_sync.success, first_sync.errors)
            self.assertEqual(first_sync.created, 1)
            original_event_id = client.service.events_data[0]["id"]

            moved = writer.update_dated_task(original.id, date=new_date)
            self.assertTrue(moved.success, moved.errors)
            after_move = writer.list_dated_tasks(new_date)[0]
            self.assertEqual(after_move.id, original.id)
            self.assertEqual(writer.list_dated_tasks(old_date), [])

            second_sync = service.calendar_sync_date(new_date)
            self.assertTrue(second_sync.success, second_sync.errors)
            self.assertEqual(second_sync.created, 0)
            self.assertEqual(second_sync.updated, 1)
            self.assertEqual(len(client.service.events_data), 1)
            self.assertEqual(client.service.events_data[0]["id"], original_event_id)
            self.assertIn("2026-07-13", client.service.events_data[0]["start"]["dateTime"])

            repeat_sync = service.calendar_sync_date(new_date)
            self.assertTrue(repeat_sync.success, repeat_sync.errors)
            self.assertEqual(repeat_sync.created, 0)
            self.assertEqual(len(client.service.events_data), 1)


class PreviewApplyTests(TestCase):
    def _preview(self, overwrite: bool = False) -> MonthlyPlanPreview:
        return MonthlyPlanPreview(
            month="Jul 2026", goals_processed=[], weekly_milestones=[], daily_tasks=[],
            feasibility_report={}, warnings=["Preview only: no workbook changes made"], conflicts=[],
            proposed_excel_changes=[{
                "operation": "add_dated_task", "date": TARGET.isoformat(), "title": "Approved task",
                "estimated_minutes": 30, "notes": "Planner OS generated plan",
            }],
            proposed_calendar_range={"start_date": TARGET.isoformat(), "end_date": TARGET.isoformat()},
            overwrite_existing_plan=overwrite,
        )

    def test_preview_persists_and_apply_accepts_id_after_reload(self) -> None:
        with TemporaryDirectory() as directory:
            tmp = Path(directory)
            _, engine, writer, _, rules = services(tmp)
            preview_dir = tmp / "previews"
            preview = self._preview()
            first = PlanningCommandService(engine, rules, writer, preview_dir)
            first._remember(preview)
            self.assertEqual(writer.list_dated_tasks(TARGET), [])

            second = PlanningCommandService(engine, rules, writer, preview_dir)
            result = second.apply_month_plan(preview.preview_id)
            self.assertTrue(result.success, result.errors)
            self.assertTrue(result.backup_path)
            self.assertEqual([task.title for task in writer.list_dated_tasks(TARGET)], ["Approved task"])

    def test_apply_accepts_preview_object_and_creates_backup(self) -> None:
        with TemporaryDirectory() as directory:
            tmp = Path(directory)
            _, engine, writer, _, rules = services(tmp)
            service = PlanningCommandService(engine, rules, writer, tmp / "previews")
            result = service.apply_month_plan(self._preview())
            self.assertTrue(result.success, result.errors)
            self.assertTrue(Path(result.backup_path).exists())

    def test_stored_planning_preview_rejects_changed_workbook(self) -> None:
        with TemporaryDirectory() as directory:
            tmp = Path(directory)
            planner_path, engine, writer, _, rules = services(tmp)
            preview_dir = tmp / "previews"
            preview = self._preview()
            service = PlanningCommandService(engine, rules, writer, preview_dir)
            service._remember(preview)
            book = load_workbook(planner_path)
            try:
                book["Jul 2026"]["L10"] = "Changed after preview"
                book.save(planner_path)
            finally:
                book.close()

            with self.assertRaisesRegex(ValueError, "(?i)preview is stale"):
                service.apply_month_plan(preview.preview_id)


class RouterTests(TestCase):
    def test_dated_intents_and_ambiguity(self) -> None:
        today = date(2026, 7, 10)
        plumber = parse_common_intent("call plumber tomorrow", today=today)
        itr = parse_common_intent("file my ITR tomorrow", today=today)
        split = parse_common_intent("sort my room half today and half tomorrow afternoon", today=today)
        ambiguous = parse_common_intent("sync week", today=today)
        self.assertEqual(plumber.command_type, "add_dated_task")
        self.assertEqual(itr.command_type, "add_dated_task")
        self.assertEqual(plumber.payload["date"], today + timedelta(days=1))
        self.assertEqual(len(split.payload["tasks"]), 2)
        self.assertTrue(ambiguous.requires_confirmation)

    def test_next_week_is_not_current_and_writes_require_confirmation(self) -> None:
        next_week = parse_common_intent("sync next week", today=date(2026, 7, 10))
        self.assertEqual(next_week.command_type, "calendar_sync_next_week")
        with TemporaryDirectory() as directory:
            _, engine, writer, _, rules = services(Path(directory))
            planning = Mock()
            calendar = Mock()
            checkin = Mock()
            manager = Mock()
            router = PlannerCommandRouter(planning, writer, manager, calendar, checkin)
            command = parse_common_intent("call plumber tomorrow", today=date(2026, 7, 10))
            refused = router.route_command(command)
            self.assertFalse(refused.success)
            self.assertTrue(refused.requires_confirmation)
            self.assertEqual(writer.list_dated_tasks(date(2026, 7, 11)), [])
            confirmed = router.route_command(
                PlannerCommand(**{**command.__dict__, "requires_confirmation": False})
            )
            self.assertTrue(confirmed.success, confirmed.errors)


class CurrentTimeAndCheckinTests(TestCase):
    def test_from_now_moves_past_flexible_blocks_and_passes_bounded_sessions(self) -> None:
        from planner_engine.current_time import CurrentTimePlanner
        from planner_engine.models import BoundedSessionRequest

        now = datetime(2026, 7, 11, 15, 7)
        scheduler = Mock()
        scheduler.plan_day.return_value = DailyPlan(
            TARGET,
            [ScheduledBlock("German", datetime(2026, 7, 11, 9), datetime(2026, 7, 11, 10), "german", "test")],
            [],
        )
        rules = RulesEngine(Path(__file__).resolve().parents[1] / "config" / "rules.yaml")
        bounded = BoundedSessionRequest("Gym", "gym", now, now + timedelta(hours=3), 2, 60, 15, True)
        plan = CurrentTimePlanner(rules, scheduler).plan_today_from_now(MonthPlan("Jul 2026", "Jul 2026", [], []), now, [bounded])
        self.assertTrue(all(block.start.replace(tzinfo=None) >= now for block in plan.blocks if not block.is_fixed))
        self.assertEqual(scheduler.plan_day.call_args.kwargs["bounded_sessions"], [bounded])

    def test_checkin_includes_dated_tasks_and_completed_recurring_counts(self) -> None:
        with TemporaryDirectory() as directory:
            _, engine, writer, progress, _ = services(Path(directory))
            self.assertTrue(writer.add_dated_task(TARGET, "German practice", 30, category="German").success)
            task = writer.list_dated_tasks(TARGET)[0]
            self.assertTrue(writer.update_dated_task(task.id, status="Done").success)
            self.assertTrue(writer.add_dated_task(TARGET, "Gym session", 60, category="Gym").success)
            gym = [item for item in writer.list_dated_tasks(TARGET) if item.title == "Gym session"][0]
            self.assertTrue(writer.update_dated_task(gym.id, status="Done").success)
            report = DailyCheckInService(engine, progress).generate_daily_checkin(TARGET)
            self.assertIn("German practice", report.planned_tasks)
            self.assertTrue(report.recurring_status["german_done"])
            self.assertGreaterEqual(report.recurring_status["gym_sessions_completed"], 1)
