from __future__ import annotations

from datetime import date
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import TestCase

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from planner_engine import (
    ExcelPlannerStore,
    PlannerEngine,
    ProgressEngine,
    RulesEngine,
    SchedulerEngine,
    Writer,
)
from planner_engine.decision_log import DecisionLog


def create_writer_workbook(path: Path) -> None:
    """Create a representative workbook for semantic writer tests."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Jul 2026"

    sheet["B8"] = "▸  MONTHLY GOALS"
    sheet["A9"] = "#"
    sheet["B9"] = "GOAL / TASK"
    sheet["C9"] = "CATEGORY"
    sheet["D9"] = "PRIORITY"
    sheet["E9"] = "TARGET WEEK"
    sheet["H9"] = "STATUS"
    sheet["L9"] = "NOTES"
    sheet.merge_cells("E9:G9")
    sheet.merge_cells("H9:K9")
    for row in range(10, 18):
        sheet.merge_cells(f"E{row}:G{row}")
        sheet.merge_cells(f"H{row}:K{row}")

    sheet["B10"] = "Existing Goal"
    sheet["C10"] = "Learning"
    sheet["D10"] = "High"
    sheet["E10"] = "W2-W5"
    sheet["H10"] = "In Progress"
    sheet["L10"] = "Goal note"

    sheet["B20"] = "WEEK 1  ·  29 Jun – 05 Jul 2026"
    sheet["A21"] = "#"
    sheet["B21"] = "TASK / GOAL"
    sheet["C21"] = "CATEGORY"
    sheet["K21"] = "STATUS"
    sheet["L21"] = "NOTES"
    sheet["B22"] = "  ↳ From monthly goals (fill in tasks planned for this week)"

    sheet["B35"] = "WEEK 2  ·  06 Jul – 12 Jul 2026"
    sheet["A36"] = "#"
    sheet["B36"] = "TASK / GOAL"
    sheet["C36"] = "CATEGORY"
    sheet["K36"] = "STATUS"
    sheet["L36"] = "NOTES"
    sheet["B37"] = "  ↳ From monthly goals (fill in tasks planned for this week)"
    sheet["B38"] = "Existing Task"
    sheet["C38"] = "Learning"
    sheet["K38"] = "Not Started"
    sheet["L38"] = "Keep this note"
    sheet["A48"] = " "
    sheet["M1"] = '=COUNTIF(K38:K48,"Done")'

    sheet["K38"].font = Font(bold=True, color="FF0000")
    sheet["K38"].fill = PatternFill(
        fill_type="solid",
        start_color="FFFF00",
        end_color="FFFF00",
    )

    status_validation = DataValidation(
        type="list",
        formula1='"Done,In Progress,To Do,Blocked,Skipped"',
    )
    status_validation.add("H10:H17")
    status_validation.add("K23:K33")
    status_validation.add("K38:K48")
    sheet.add_data_validation(status_validation)

    priority_validation = DataValidation(type="list", formula1='"High,Medium,Low"')
    priority_validation.add("D10:D17")
    sheet.add_data_validation(priority_validation)

    workbook.save(path)
    workbook.close()


def writer_for(planner_path: Path, backup_dir: Path) -> tuple[Writer, ProgressEngine]:
    """Create a semantic writer and progress engine for tests."""

    progress = ProgressEngine()
    engine = PlannerEngine(ExcelPlannerStore(planner_path, backup_dir))
    decision_log = DecisionLog(backup_dir.parent / "decision_log.jsonl")
    return Writer(engine, progress, decision_log=decision_log), progress


class SemanticWriterTests(TestCase):
    """Tests for safe semantic planner writes."""

    def test_add_goal(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            backup_dir = tmp_path / "backups"
            create_writer_workbook(planner_path)
            writer, _ = writer_for(planner_path, backup_dir)

            result = writer.add_monthly_goal(
                "Jul 2026",
                "New Goal",
                category="Planning",
                priority="Medium",
                target_week="W3",
                status="Not Started",
                notes="New note",
            )

            self.assertTrue(result.success)
            self.assertEqual(len(list(backup_dir.glob("*.xlsx"))), 1)
            workbook = load_workbook(planner_path)
            try:
                sheet = workbook["Jul 2026"]
                self.assertEqual(sheet["B11"].value, "New Goal")
                self.assertEqual(sheet["C11"].value, "Planning")
                self.assertEqual(sheet["D11"].value, "Medium")
                self.assertEqual(sheet["E11"].value, "W3")
                self.assertEqual(sheet["H11"].value, "Not Started")
                self.assertEqual(sheet["L11"].value, "New note")
            finally:
                workbook.close()

    def test_add_task(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            backup_dir = tmp_path / "backups"
            create_writer_workbook(planner_path)
            writer, _ = writer_for(planner_path, backup_dir)

            result = writer.add_weekly_task(
                "Jul 2026",
                2,
                "New Weekly Task",
                category="Study",
                status="To Do",
                notes="Task note",
            )

            self.assertTrue(result.success)
            self.assertEqual(len(list(backup_dir.glob("*.xlsx"))), 1)
            workbook = load_workbook(planner_path)
            try:
                sheet = workbook["Jul 2026"]
                self.assertEqual(sheet["B39"].value, "New Weekly Task")
                self.assertEqual(sheet["C39"].value, "Study")
                self.assertEqual(sheet["K39"].value, "To Do")
                self.assertEqual(sheet["L39"].value, "Task note")
            finally:
                workbook.close()

    def test_full_week_expands_with_formatting_validation_and_reader_support(
        self,
    ) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            backup_dir = tmp_path / "backups"
            create_writer_workbook(planner_path)
            workbook = load_workbook(planner_path)
            try:
                sheet = workbook["Jul 2026"]
                for row in range(23, 35):
                    sheet.cell(row, 2).value = f"Existing Week 1 Task {row}"
                    sheet.cell(row, 3).value = "Learning"
                    sheet.cell(row, 11).value = "Not Started"
                sheet["B34"].font = Font(bold=True, color="001122")
                sheet["B34"].fill = PatternFill(
                    fill_type="solid",
                    start_color="AABBCC",
                    end_color="AABBCC",
                )
                sheet["M34"] = "=LEN(B34)"
                sheet["N1"] = '=COUNTIF(K10:K200,"Done")'
                sheet.merge_cells("D34:J34")
                sheet.data_validations.dataValidation[0].add("K34")

                category_validation = DataValidation(
                    type="list",
                    formula1='"Work,Learning,Health,Personal"',
                )
                category_validation.add("C23:C34")
                category_validation.add("C38:C48")
                sheet.add_data_validation(category_validation)
                workbook.save(planner_path)
            finally:
                workbook.close()

            writer, _ = writer_for(planner_path, backup_dir)
            result = writer.add_weekly_task(
                "Jul 2026",
                1,
                "Expanded Weekly Task",
                category="Health",
                status="To Do",
                notes="Inserted row",
            )

            self.assertTrue(result.success, result.errors)
            workbook = load_workbook(planner_path)
            try:
                sheet = workbook["Jul 2026"]
                self.assertEqual(sheet["B35"].value, "Expanded Weekly Task")
                self.assertEqual(sheet["B36"].value, "WEEK 2  ·  06 Jul – 12 Jul 2026")
                self.assertEqual(sheet["B35"].style_id, sheet["B34"].style_id)
                self.assertEqual(sheet["M35"].value, "=LEN(B35)")
                self.assertEqual(sheet["M1"].value, '=COUNTIF(K39:K49,"Done")')
                self.assertEqual(sheet["N1"].value, '=COUNTIF(K10:K201,"Done")')
                self.assertIn("D35:J35", {str(cell_range) for cell_range in sheet.merged_cells.ranges})
                validation_ranges = [
                    cell_range
                    for validation in sheet.data_validations.dataValidation
                    for cell_range in validation.ranges.ranges
                ]
                self.assertTrue(
                    any("C35" in cell_range for cell_range in validation_ranges)
                )
                self.assertTrue(
                    any("K35" in cell_range for cell_range in validation_ranges)
                )
            finally:
                workbook.close()

            engine = PlannerEngine(ExcelPlannerStore(planner_path, backup_dir))
            month_plan = engine.get_month_plan("Jul 2026")
            self.assertEqual(month_plan.week_sections[1].heading_row, 36)
            self.assertIn(
                "Expanded Weekly Task",
                [
                    task.name
                    for section in month_plan.week_sections
                    for task in section.tasks
                ],
            )
            schedule = SchedulerEngine(
                RulesEngine(Path("config/rules.yaml")),
                decision_log=DecisionLog(tmp_path / "scheduler_log.jsonl"),
            ).plan_week(month_plan, date(2026, 7, 6))
            scheduler_items = {
                block.title for day in schedule.days for block in day.blocks
            } | {conflict.item for conflict in schedule.conflicts}
            self.assertIn("Expanded Weekly Task", scheduler_items)

    def test_update_task(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            backup_dir = tmp_path / "backups"
            create_writer_workbook(planner_path)
            writer, _ = writer_for(planner_path, backup_dir)

            result = writer.update_task(
                "Jul 2026",
                "Existing Task",
                category="Study",
                status="In Progress",
                notes="Updated note",
            )

            self.assertTrue(result.success)
            self.assertEqual(len(list(backup_dir.glob("*.xlsx"))), 1)
            workbook = load_workbook(planner_path)
            try:
                sheet = workbook["Jul 2026"]
                self.assertEqual(sheet["B38"].value, "Existing Task")
                self.assertEqual(sheet["C38"].value, "Study")
                self.assertEqual(sheet["K38"].value, "In Progress")
                self.assertEqual(sheet["L38"].value, "Updated note")
            finally:
                workbook.close()

    def test_complete_task_updates_progress_engine(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            backup_dir = tmp_path / "backups"
            create_writer_workbook(planner_path)
            writer, progress = writer_for(planner_path, backup_dir)

            result = writer.complete_task(
                "Jul 2026",
                "Existing Task",
                completion_date=date(2026, 7, 10),
            )

            self.assertTrue(result.success)
            self.assertEqual(len(list(backup_dir.glob("*.xlsx"))), 1)
            self.assertIsNotNone(result.progress_execution)
            self.assertEqual(len(progress.executions), 1)
            self.assertEqual(progress.executions[0].task_name, "Existing Task")
            workbook = load_workbook(planner_path)
            try:
                self.assertEqual(workbook["Jul 2026"]["K38"].value, "Done")
            finally:
                workbook.close()

    def test_move_task(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            backup_dir = tmp_path / "backups"
            create_writer_workbook(planner_path)
            writer, _ = writer_for(planner_path, backup_dir)

            result = writer.move_task("Jul 2026", "Existing Task", 1)

            self.assertTrue(result.success)
            self.assertEqual(len(list(backup_dir.glob("*.xlsx"))), 1)
            workbook = load_workbook(planner_path)
            try:
                sheet = workbook["Jul 2026"]
                self.assertEqual(sheet["B23"].value, "Existing Task")
                self.assertEqual(sheet["C23"].value, "Learning")
                self.assertEqual(sheet["K23"].value, "Not Started")
                self.assertEqual(sheet["L23"].value, "Keep this note")
                self.assertIsNone(sheet["B38"].value)
                self.assertIsNone(sheet["C38"].value)
                self.assertIsNone(sheet["K38"].value)
                self.assertIsNone(sheet["L38"].value)
            finally:
                workbook.close()

    def test_delete_task(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            backup_dir = tmp_path / "backups"
            create_writer_workbook(planner_path)
            writer, _ = writer_for(planner_path, backup_dir)

            result = writer.delete_task("Jul 2026", "Existing Task")

            self.assertTrue(result.success)
            self.assertEqual(len(list(backup_dir.glob("*.xlsx"))), 1)
            workbook = load_workbook(planner_path)
            try:
                sheet = workbook["Jul 2026"]
                self.assertIsNone(sheet["B38"].value)
                self.assertIsNone(sheet["C38"].value)
                self.assertIsNone(sheet["K38"].value)
                self.assertIsNone(sheet["L38"].value)
            finally:
                workbook.close()

    def test_duplicate_task(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            backup_dir = tmp_path / "backups"
            create_writer_workbook(planner_path)
            writer, _ = writer_for(planner_path, backup_dir)

            result = writer.add_weekly_task("Jul 2026", 2, "Existing Task")

            self.assertFalse(result.success)
            self.assertEqual(result.errors, ("Duplicate task in week 2: Existing Task",))
            self.assertEqual(len(list(backup_dir.glob("*.xlsx"))), 0)

    def test_rollback_safety(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            backup_dir = tmp_path / "backups"
            create_writer_workbook(planner_path)
            progress = ProgressEngine()
            engine = PlannerEngine(ExcelPlannerStore(planner_path, backup_dir))
            writer = Writer(
                engine,
                progress,
                decision_log=DecisionLog(tmp_path / "decision_log.jsonl"),
            )
            original_write = engine._write_cells_without_backup

            def fail_after_write(updates: list) -> None:
                original_write(updates)
                raise RuntimeError("forced failure")

            engine._write_cells_without_backup = fail_after_write  # type: ignore[method-assign]

            result = writer.update_task("Jul 2026", "Existing Task", status="Done")

            self.assertFalse(result.success)
            self.assertEqual(result.errors, ("forced failure",))
            self.assertEqual(len(list(backup_dir.glob("*.xlsx"))), 1)
            workbook = load_workbook(planner_path)
            try:
                self.assertEqual(workbook["Jul 2026"]["K38"].value, "Not Started")
            finally:
                workbook.close()

    def test_formatting_layout_validation_and_formulas_are_preserved(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            backup_dir = tmp_path / "backups"
            create_writer_workbook(planner_path)
            writer, _ = writer_for(planner_path, backup_dir)

            result = writer.update_task("Jul 2026", "Existing Task", status="Done")

            self.assertTrue(result.success)
            workbook = load_workbook(planner_path, data_only=False)
            try:
                sheet = workbook["Jul 2026"]
                cell = sheet["K38"]
                self.assertTrue(cell.font.bold)
                self.assertEqual(cell.font.color.rgb, "00FF0000")
                self.assertEqual(cell.fill.fill_type, "solid")
                self.assertEqual(cell.fill.start_color.rgb, "00FFFF00")
                self.assertIn("H9:K9", [str(cell_range) for cell_range in sheet.merged_cells.ranges])
                self.assertEqual(sheet["M1"].value, '=COUNTIF(K38:K48,"Done")')
                self.assertGreater(len(sheet.data_validations.dataValidation), 0)
            finally:
                workbook.close()
