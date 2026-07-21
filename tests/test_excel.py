from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import TestCase

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from planner_engine import ExcelPlannerStore


def create_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Plan"
    worksheet["A1"] = "Original"
    worksheet["A1"].font = Font(bold=True, color="FF0000")
    worksheet["A1"].fill = PatternFill(
        fill_type="solid",
        start_color="FFFF00",
        end_color="FFFF00",
    )
    workbook.save(path)
    workbook.close()


class ExcelPlannerStoreTests(TestCase):
    """Tests for Excel workbook storage behavior."""

    def test_read_cell(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            create_workbook(planner_path)
            store = ExcelPlannerStore(planner_path, tmp_path / "backups")

            self.assertEqual(store.read_cell("Plan", "A1"), "Original")

    def test_write_cell(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            create_workbook(planner_path)
            store = ExcelPlannerStore(planner_path, tmp_path / "backups")

            store.write_cell("Plan", "A1", "Updated")

            self.assertEqual(store.read_cell("Plan", "A1"), "Updated")

    def test_backup_creation(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            backup_dir = tmp_path / "backups"
            create_workbook(planner_path)
            store = ExcelPlannerStore(planner_path, backup_dir)

            backup_path = store.create_backup()

            self.assertTrue(backup_path.exists())
            self.assertEqual(backup_path.parent, backup_dir)
            backup_workbook = load_workbook(backup_path)
            try:
                self.assertEqual(backup_workbook["Plan"]["A1"].value, "Original")
            finally:
                backup_workbook.close()

    def test_read_workbook_is_parsed_once_and_reused(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            create_workbook(planner_path)
            store = ExcelPlannerStore(planner_path, tmp_path / "backups")

            first = store.load_workbook_read_only()
            second = store.load_workbook_read_only()
            # Same object back: the file was parsed once, not twice.
            self.assertIs(first, second)

    def test_write_invalidates_read_cache(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            create_workbook(planner_path)
            store = ExcelPlannerStore(planner_path, tmp_path / "backups")

            # Warm the cached reader path and confirm it sees the original value.
            cached = store.load_workbook_read_only()
            self.assertEqual(cached["Plan"]["A1"].value, "Original")

            store.write_cell("Plan", "A1", "Updated")

            # After a write the cached reader must reload — a different object
            # reflecting the new value, never the stale parse.
            refreshed = store.load_workbook_read_only()
            self.assertIsNot(refreshed, cached)
            self.assertEqual(refreshed["Plan"]["A1"].value, "Updated")

    def test_restore_backup_invalidates_read_cache(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            store = ExcelPlannerStore(planner_path, tmp_path / "backups")
            create_workbook(planner_path)

            backup_path = store.create_backup()  # snapshot with "Original"
            store.write_cell("Plan", "A1", "Updated")
            self.assertEqual(store.read_cell("Plan", "A1"), "Updated")

            store.restore_backup(backup_path)
            # Restore replaces the file underneath the cache; reads must refresh.
            self.assertEqual(store.read_cell("Plan", "A1"), "Original")

    def test_rapid_backups_have_unique_filenames(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            backup_dir = tmp_path / "backups"
            create_workbook(planner_path)
            store = ExcelPlannerStore(planner_path, backup_dir)

            backup_paths = [store.create_backup() for _ in range(5)]

            self.assertEqual(len(set(backup_paths)), 5)
            self.assertEqual(len(list(backup_dir.glob("*.xlsx"))), 5)
            for backup_path in backup_paths:
                self.assertTrue(backup_path.exists())

    def test_original_formatting_is_preserved_after_write(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            create_workbook(planner_path)
            store = ExcelPlannerStore(planner_path, tmp_path / "backups")

            store.write_cell("Plan", "A1", "Updated")

            workbook = load_workbook(planner_path)
            try:
                cell = workbook["Plan"]["A1"]
                self.assertEqual(cell.value, "Updated")
                self.assertTrue(cell.font.bold)
                self.assertEqual(cell.font.color.rgb, "00FF0000")
                self.assertEqual(cell.fill.fill_type, "solid")
                self.assertEqual(cell.fill.start_color.rgb, "00FFFF00")
            finally:
                workbook.close()
