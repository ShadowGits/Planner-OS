import json
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import TestCase

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from planner_engine import ExcelPlannerStore, PlannerEngine, PlannerImporter


def create_import_workbook(path: Path) -> None:
    """Create a workbook representative enough for importer tests."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Jul 2026"

    sheet["B8"] = "▸  MONTHLY GOALS"
    sheet["A9"] = "#"
    sheet["B9"] = "GOAL / TASK"
    sheet["C9"] = "CATEGORY"
    sheet["D9"] = "PRIORITY"
    sheet["E9"] = "TARGET WEEK"
    sheet["H9"] = "STATUS"
    sheet["L9"] = "NOTES"
    sheet.merge_cells("E9:G9")
    sheet.merge_cells("H9:K9")
    for row in range(10, 18):
        sheet.merge_cells(f"E{row}:G{row}")
        sheet.merge_cells(f"H{row}:K{row}")

    sheet["B20"] = "WEEK 1  ·  29 Jun – 05 Jul 2026"
    sheet["A21"] = "#"
    sheet["B21"] = "TASK / GOAL"
    sheet["C21"] = "CATEGORY"
    sheet["K21"] = "STATUS"
    sheet["L21"] = "NOTES"
    sheet["B22"] = "  ↳ From monthly goals (fill in tasks planned for this week)"

    sheet["B35"] = "WEEK 2  ·  06 Jul – 12 Jul 2026"
    sheet["A36"] = "#"
    sheet["B36"] = "TASK / GOAL"
    sheet["C36"] = "CATEGORY"
    sheet["K36"] = "STATUS"
    sheet["L36"] = "NOTES"
    sheet["B37"] = "  ↳ From monthly goals (fill in tasks planned for this week)"
    sheet["A48"] = " "

    status_validation = DataValidation(
        type="list",
        formula1='"Done,In Progress,To Do,Blocked,Skipped"',
    )
    status_validation.add("H10:H17")
    status_validation.add("K23:K33")
    status_validation.add("K38:K48")
    sheet.add_data_validation(status_validation)

    priority_validation = DataValidation(type="list", formula1='"High,Medium,Low"')
    priority_validation.add("D10:D17")
    sheet.add_data_validation(priority_validation)

    category_validation = DataValidation(
        type="list",
        formula1='"Work,Learning,Health,Personal,Finance,German,Reading,Other"',
    )
    category_validation.add("C10:C17")
    category_validation.add("C23:C33")
    category_validation.add("C38:C48")
    sheet.add_data_validation(category_validation)

    target_week_validation = DataValidation(
        type="list",
        formula1='"Week 1,Week 2,Week 3,Week 4,Week 5,All Month"',
    )
    target_week_validation.add("E10:E17")
    sheet.add_data_validation(target_week_validation)

    workbook.save(path)
    workbook.close()


def valid_document() -> dict:
    """Return a valid importer JSON-like document."""

    return {
        "month": "Jul 2026",
        "monthly_goals": [
            {
                "goal": "Complete A1 German module",
                "category": "Learning",
                "priority": "High",
                "target_week": "W2-W5",
                "status": "In Progress",
                "notes": "",
            }
        ],
        "weeks": [
            {
                "week": 2,
                "tasks": [
                    {
                        "task": "Babbel Unit 3",
                        "category": "Learning",
                        "status": "Not Started",
                        "notes": "",
                    }
                ],
            }
        ],
    }


def importer_for(planner_path: Path, backup_dir: Path) -> PlannerImporter:
    """Create an importer for a test workbook."""

    engine = PlannerEngine(ExcelPlannerStore(planner_path, backup_dir))
    return PlannerImporter(engine)


class PlannerImporterTests(TestCase):
    """Tests for structured JSON planner imports."""

    def test_public_monthly_goal_append_creates_one_backup(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            backup_dir = tmp_path / "backups"
            create_import_workbook(planner_path)
            engine = PlannerEngine(ExcelPlannerStore(planner_path, backup_dir))

            imported = engine.append_monthly_goals(
                "Jul 2026",
                valid_document()["monthly_goals"],
            )

            self.assertEqual(imported, 1)
            self.assertEqual(len(list(backup_dir.glob("*.xlsx"))), 1)

    def test_public_weekly_task_append_creates_one_backup(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            backup_dir = tmp_path / "backups"
            create_import_workbook(planner_path)
            engine = PlannerEngine(ExcelPlannerStore(planner_path, backup_dir))

            imported = engine.append_weekly_tasks(
                "Jul 2026",
                [{"week": 2, **valid_document()["weeks"][0]["tasks"][0]}],
            )

            self.assertEqual(imported, 1)
            self.assertEqual(len(list(backup_dir.glob("*.xlsx"))), 1)

    def test_valid_import(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            backup_dir = tmp_path / "backups"
            create_import_workbook(planner_path)

            result = importer_for(planner_path, backup_dir).import_document(
                valid_document()
            )

            self.assertTrue(result.success)
            self.assertEqual(result.goals_imported, 1)
            self.assertEqual(result.tasks_imported, 1)
            workbook = load_workbook(planner_path)
            try:
                sheet = workbook["Jul 2026"]
                self.assertEqual(sheet["B10"].value, "Complete A1 German module")
                self.assertEqual(sheet["C10"].value, "Learning")
                self.assertEqual(sheet["D10"].value, "High")
                self.assertEqual(sheet["E10"].value, "W2-W5")
                self.assertEqual(sheet["H10"].value, "In Progress")
                self.assertEqual(sheet["B38"].value, "Babbel Unit 3")
                self.assertEqual(sheet["C38"].value, "Learning")
                self.assertEqual(sheet["K38"].value, "Not Started")
            finally:
                workbook.close()

    def test_invalid_month(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            create_import_workbook(planner_path)
            document = valid_document()
            document["month"] = "Aug 2026"

            result = importer_for(planner_path, tmp_path / "backups").import_document(
                document
            )

            self.assertFalse(result.success)
            self.assertEqual(result.validation_errors, ["Unknown planner month: Aug 2026"])

    def test_invalid_status(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            create_import_workbook(planner_path)
            document = valid_document()
            document["weeks"][0]["tasks"][0]["status"] = "Almost Done"

            result = importer_for(planner_path, tmp_path / "backups").import_document(
                document
            )

            self.assertFalse(result.success)
            self.assertIn(
                "Invalid status for weekly task 'Babbel Unit 3': Almost Done",
                result.validation_errors,
            )

    def test_invalid_priority(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            create_import_workbook(planner_path)
            document = valid_document()
            document["monthly_goals"][0]["priority"] = "Urgent"

            result = importer_for(planner_path, tmp_path / "backups").import_document(
                document
            )

            self.assertFalse(result.success)
            self.assertIn(
                "Invalid priority for monthly goal 'Complete A1 German module': Urgent",
                result.validation_errors,
            )

    def test_malformed_json(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            input_path = tmp_path / "bad.json"
            create_import_workbook(planner_path)
            input_path.write_text('{"month": "Jul 2026"', encoding="utf-8")

            result = importer_for(planner_path, tmp_path / "backups").import_file(
                input_path
            )

            self.assertFalse(result.success)
            self.assertEqual(result.validation_errors[0], "Malformed JSON: Expecting ',' delimiter")

    def test_duplicate_task_handling(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            backup_dir = tmp_path / "backups"
            create_import_workbook(planner_path)
            document = valid_document()
            document["weeks"][0]["tasks"].append(dict(document["weeks"][0]["tasks"][0]))

            result = importer_for(planner_path, backup_dir).import_document(document)

            self.assertTrue(result.success)
            self.assertEqual(result.tasks_imported, 1)
            self.assertEqual(
                result.skipped_items,
                ["Duplicate weekly task skipped: W2 Babbel Unit 3"],
            )

    def test_rollback_on_failure(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            backup_dir = tmp_path / "backups"
            create_import_workbook(planner_path)
            engine = PlannerEngine(ExcelPlannerStore(planner_path, backup_dir))

            def fail_weekly_tasks(month: str, week_tasks: list[dict]) -> int:
                raise RuntimeError("forced failure")

            engine._append_weekly_tasks_without_backup = fail_weekly_tasks  # type: ignore[method-assign]
            result = PlannerImporter(engine).import_document(valid_document())

            self.assertFalse(result.success)
            self.assertEqual(len(list(backup_dir.glob("*.xlsx"))), 1)
            workbook = load_workbook(planner_path)
            try:
                self.assertIsNone(workbook["Jul 2026"]["B10"].value)
                self.assertIsNone(workbook["Jul 2026"]["B38"].value)
            finally:
                workbook.close()

    def test_exactly_one_backup_created(self) -> None:
        with TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            planner_path = tmp_path / "planner.xlsx"
            backup_dir = tmp_path / "backups"
            create_import_workbook(planner_path)

            result = importer_for(planner_path, backup_dir).import_document(
                valid_document()
            )

            self.assertTrue(result.success)
            self.assertEqual(len(list(backup_dir.glob("*.xlsx"))), 1)
            self.assertIsNotNone(result.backup_path)
