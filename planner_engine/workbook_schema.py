"""Template-aware workbook capability and schema inspection."""

from __future__ import annotations

from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True)
class WorkbookCapabilities:
    supports_long_term_goals: bool = False
    supports_annual_goals: bool = False
    supports_quarterly_goals: bool = False
    supports_monthly_goals: bool = False
    supports_weekly_tasks: bool = False
    supports_dated_tasks: bool = False
    supports_start_time: bool = False
    supports_end_time: bool = False
    supports_duration: bool = False
    supports_hard_time: bool = False
    supports_progress_tracking: bool = False

    def to_dict(self):
        return asdict(self)


@dataclass(frozen=True)
class WorkbookSchemaIssue:
    sheet: str
    section: str
    row: int | None
    column: int | None
    expected: str
    actual: str

    def to_dict(self):
        return asdict(self)


class WorkbookSchemaInspector:
    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)

    def inspect(self) -> tuple[WorkbookCapabilities, list[WorkbookSchemaIssue]]:
        workbook = load_workbook(self.path, read_only=False, data_only=False)
        try:
            labels = []
            date_columns = False
            time_labels = set()
            progress = False
            week_headers = False
            issues = []
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    values = [str(cell.value).strip() if cell.value is not None else "" for cell in row]
                    normalized = " | ".join(values).casefold()
                    labels.append(normalized)
                    if any(value.upper().startswith("WEEK ") for value in values):
                        week_headers = True
                    if "status" in normalized:
                        progress = True
                    for key in ("start time", "end time", "duration", "hard time", "hard/flexible"):
                        if key in normalized:
                            time_labels.add(key)
                    if any(any(month in value.upper() for month in (" JAN", " FEB", " MAR", " APR", " MAY", " JUN", " JUL", " AUG", " SEP", " OCT", " NOV", " DEC")) and any(char.isdigit() for char in value) for value in values):
                        date_columns = True
            text = "\n".join(labels)
            monthly = "monthly goals" in text
            weekly = week_headers and ("task / goal" in text or "goal / task" in text)
            dated = weekly and date_columns
            if not monthly:
                issues.append(WorkbookSchemaIssue("*", "monthly_goals", None, None, "MONTHLY GOALS section", "not found"))
            if not weekly:
                issues.append(WorkbookSchemaIssue("*", "weekly_tasks", None, None, "WEEK heading and TASK / GOAL header", "not found"))
            return WorkbookCapabilities(
                supports_long_term_goals="long-term goals" in text or "long term goals" in text,
                supports_annual_goals="annual goals" in text,
                supports_quarterly_goals="quarterly goals" in text,
                supports_monthly_goals=monthly,
                supports_weekly_tasks=weekly,
                supports_dated_tasks=dated,
                supports_start_time="start time" in time_labels or dated,
                supports_end_time="end time" in time_labels or dated,
                supports_duration="duration" in time_labels or dated,
                supports_hard_time=bool({"hard time", "hard/flexible"} & time_labels) or dated,
                supports_progress_tracking=progress,
            ), issues
        finally:
            workbook.close()
