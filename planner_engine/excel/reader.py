"""Semantic read support for Excel planner workbooks."""

from __future__ import annotations

from datetime import date, datetime, timedelta
import re
from typing import Any

from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from planner_engine.excel.layout import PlannerWorkbookError
from planner_engine.models import (
    MonthPlan,
    DatedTask,
    MonthlyGoal,
    PlannerTask,
    Priority,
    TaskStatus,
    WeekSection,
)


class ReaderMixin:
    """Parse planner concepts from workbook sheets."""

    def list_months(self) -> list[str]:
        """List monthly planner sheet names."""

        workbook = self.load_workbook_read_only()
        try:
            return [
                sheet_name
                for sheet_name in workbook.sheetnames
                if self._looks_like_month_sheet(sheet_name)
            ]
        finally:
            workbook.close()

    def read_month_plan(self, month: str) -> MonthPlan:
        """Read a full parsed month plan without modifying the workbook."""

        workbook = self.load_workbook_read_only()
        try:
            worksheet = self._get_month_worksheet(workbook, month)
            return MonthPlan(
                month=worksheet.title,
                sheet_name=worksheet.title,
                monthly_goals=self._parse_monthly_goals(worksheet),
                week_sections=self._parse_week_sections(worksheet),
            )
        finally:
            workbook.close()

    def read_monthly_goals(self, month: str) -> list[MonthlyGoal]:
        """Read monthly goals from a month sheet."""

        return self.read_month_plan(month).monthly_goals

    def read_week_sections(self, month: str) -> list[WeekSection]:
        """Read weekly sections from a month sheet."""

        return self.read_month_plan(month).week_sections

    def read_dated_tasks(
        self,
        month: str,
        target_date: date | None = None,
    ) -> list[DatedTask]:
        """Read tasks assigned to exact day columns in a month sheet."""

        workbook = self.load_workbook_read_only()
        try:
            worksheet = self._get_month_worksheet(workbook, month)
            tasks: list[DatedTask] = []
            for section in self._parse_week_sections(worksheet):
                week_start, week_end = self._week_title_range(section.title, month)
                for task in section.tasks:
                    for offset, task_date in enumerate(
                        week_start + timedelta(days=index)
                        for index in range((week_end - week_start).days + 1)
                    ):
                        if offset >= 7 or (target_date is not None and task_date != target_date):
                            continue
                        column = 4 + offset
                        raw_block = self._string_or_none(
                            worksheet.cell(task.row_number, column).value
                        )
                        if not raw_block:
                            continue
                        minutes, daypart, start_time, end_time, hard_time = (
                            self._parse_dated_block(raw_block)
                        )
                        task_id = f"{worksheet.title}!{task.row_number}@{task_date.isoformat()}"
                        tasks.append(
                            DatedTask(
                                id=task_id,
                                date=task_date,
                                title=task.name,
                                estimated_minutes=minutes,
                                preferred_daypart=daypart,
                                start_time=start_time,
                                end_time=end_time,
                                hard_time=hard_time,
                                category=task.category,
                                status=task.status,
                                notes=task.notes,
                                sheet_name=worksheet.title,
                                row_number=task.row_number,
                                week_name=section.name,
                                cell_references={
                                    **task.cell_references,
                                    "time_block": worksheet.cell(task.row_number, column).coordinate,
                                },
                            )
                        )
            return tasks
        finally:
            workbook.close()

    def find_task(self, month: str, task_name: str) -> PlannerTask | MonthlyGoal | None:
        """Find a task or goal by case-insensitive name."""

        normalized_task_name = self._normalize_lookup(task_name)
        month_plan = self.read_month_plan(month)

        for goal in month_plan.monthly_goals:
            if self._normalize_lookup(goal.name) == normalized_task_name:
                return goal

        for week_section in month_plan.week_sections:
            for task in week_section.tasks:
                if self._normalize_lookup(task.name) == normalized_task_name:
                    return task

        return None

    def _get_month_worksheet(self, workbook: Workbook, month: str) -> Worksheet:
        """Return a month worksheet, or raise a clear parsing error."""

        requested_month = self._normalize_lookup(month)
        for sheet_name in workbook.sheetnames:
            if self._normalize_lookup(sheet_name) == requested_month:
                return workbook[sheet_name]

        raise PlannerWorkbookError(f"Unknown planner month: {month}")

    def _parse_monthly_goals(self, worksheet: Worksheet) -> list[MonthlyGoal]:
        """Parse monthly goals using heading labels."""

        header_row, end_row, header_map = self._monthly_goal_table(worksheet)
        task_column = self._required_column(header_map, ("GOAL / TASK", "TASK / GOAL"))

        goals: list[MonthlyGoal] = []
        for row_number in range(header_row + 1, end_row + 1):
            name = self._string_or_none(worksheet.cell(row_number, task_column).value)
            if not name:
                continue

            category = self._read_string(worksheet, row_number, header_map, "CATEGORY")
            raw_priority = self._read_string(
                worksheet,
                row_number,
                header_map,
                "PRIORITY",
            )
            target_week = self._read_string(
                worksheet,
                row_number,
                header_map,
                "TARGET WEEK",
            )
            raw_status = self._read_string(worksheet, row_number, header_map, "STATUS")
            notes = self._read_string(worksheet, row_number, header_map, "NOTES")

            goals.append(
                MonthlyGoal(
                    name=name,
                    category=category,
                    priority=Priority.from_value(raw_priority),
                    target_week=target_week,
                    status=TaskStatus.from_value(raw_status),
                    notes=notes,
                    sheet_name=worksheet.title,
                    row_number=row_number,
                    cell_references=self._cell_references(
                        worksheet,
                        row_number,
                        header_map,
                        {
                            "name": ("GOAL / TASK", "TASK / GOAL"),
                            "category": ("CATEGORY",),
                            "priority": ("PRIORITY",),
                            "target_week": ("TARGET WEEK",),
                            "status": ("STATUS",),
                            "notes": ("NOTES",),
                        },
                    ),
                    raw_priority=raw_priority,
                    raw_status=raw_status,
                )
            )

        return goals

    def _parse_week_sections(self, worksheet: Worksheet) -> list[WeekSection]:
        """Parse all weekly sections using WEEK labels."""

        week_heading_rows = self._find_week_heading_rows(worksheet)
        week_sections: list[WeekSection] = []

        for index, heading_row in enumerate(week_heading_rows):
            next_heading_row = (
                week_heading_rows[index + 1]
                if index + 1 < len(week_heading_rows)
                else worksheet.max_row + 1
            )
            header_row = self._find_header_row_after(
                worksheet,
                start_row=heading_row,
                required_labels=("TASK / GOAL", "CATEGORY", "STATUS", "NOTES"),
                stop_row=next_heading_row,
            )
            title = self._string_or_none(worksheet.cell(heading_row, 2).value) or ""
            week_name = title.split("·", maxsplit=1)[0].strip()
            start_row = header_row + 1
            end_row = next_heading_row - 1
            header_map = self._header_map(
                worksheet,
                header_row,
                data_start_row=start_row,
                data_end_row=end_row,
            )

            tasks = self._parse_week_tasks(
                worksheet=worksheet,
                week_name=week_name,
                week_title=title,
                header_row=header_row,
                start_row=start_row,
                end_row=end_row,
                header_map=header_map,
            )
            week_sections.append(
                WeekSection(
                    name=week_name,
                    title=title,
                    sheet_name=worksheet.title,
                    heading_row=heading_row,
                    header_row=header_row,
                    start_row=start_row,
                    end_row=end_row,
                    tasks=tasks,
                    cell_references={
                        "heading": worksheet.cell(heading_row, 2).coordinate,
                    },
                )
            )

        return week_sections

    def _parse_week_tasks(
        self,
        worksheet: Worksheet,
        week_name: str,
        week_title: str,
        header_row: int,
        start_row: int,
        end_row: int,
        header_map: dict[str, int],
    ) -> list[PlannerTask]:
        """Parse task rows in a weekly section."""

        task_column = self._required_column(header_map, ("TASK / GOAL", "GOAL / TASK"))
        tasks: list[PlannerTask] = []

        for row_number in range(start_row, end_row + 1):
            name = self._string_or_none(worksheet.cell(row_number, task_column).value)
            if not name or name.startswith("↳"):
                continue

            category = self._read_string(worksheet, row_number, header_map, "CATEGORY")
            raw_status = self._read_string(worksheet, row_number, header_map, "STATUS")
            notes = self._read_string(worksheet, row_number, header_map, "NOTES")

            tasks.append(
                PlannerTask(
                    name=name,
                    category=category,
                    priority=Priority.UNKNOWN,
                    status=TaskStatus.from_value(raw_status),
                    notes=notes,
                    sheet_name=worksheet.title,
                    row_number=row_number,
                    week_name=week_name,
                    cell_references=self._cell_references(
                        worksheet,
                        row_number,
                        header_map,
                        {
                            "name": ("TASK / GOAL", "GOAL / TASK"),
                            "category": ("CATEGORY",),
                            "status": ("STATUS",),
                            "notes": ("NOTES",),
                        },
                    ),
                    raw_status=raw_status,
                    scheduled_dates=self._scheduled_dates_for_row(
                        worksheet,
                        row_number,
                        week_title,
                    ),
                )
            )

        return tasks

    def _scheduled_dates_for_row(
        self,
        worksheet: Worksheet,
        row_number: int,
        week_title: str,
    ) -> tuple[date, ...]:
        week_start, week_end = self._week_title_range(week_title, worksheet.title)
        dates: list[date] = []
        for offset in range(min(7, (week_end - week_start).days + 1)):
            if self._string_or_none(worksheet.cell(row_number, 4 + offset).value):
                dates.append(week_start + timedelta(days=offset))
        return tuple(dates)

    def _monthly_goal_table(
        self,
        worksheet: Worksheet,
    ) -> tuple[int, int, dict[str, int]]:
        """Return monthly goal header row, end row, and header map."""

        section_row = self._find_label_row(worksheet, "MONTHLY GOALS")
        header_row = self._find_header_row_after(
            worksheet,
            start_row=section_row,
            required_labels=("GOAL / TASK", "CATEGORY", "PRIORITY", "TARGET WEEK"),
        )
        next_week_row = self._find_next_week_heading_row(
            worksheet,
            start_row=header_row + 1,
        )
        end_row = (next_week_row - 1) if next_week_row else worksheet.max_row
        header_map = self._header_map(
            worksheet,
            header_row,
            data_start_row=header_row + 1,
            data_end_row=end_row,
        )
        return header_row, end_row, header_map

    def _week_task_table(
        self,
        worksheet: Worksheet,
        week_number: int,
    ) -> tuple[int, int, dict[str, int]]:
        """Return a week task header row, end row, and header map."""

        week_heading_rows = self._find_week_heading_rows(worksheet)
        if week_number < 1 or week_number > len(week_heading_rows):
            raise PlannerWorkbookError(
                f"Unknown week {week_number} in sheet '{worksheet.title}'"
            )

        heading_row = week_heading_rows[week_number - 1]
        next_heading_row = (
            week_heading_rows[week_number]
            if week_number < len(week_heading_rows)
            else worksheet.max_row + 1
        )
        header_row = self._find_header_row_after(
            worksheet,
            start_row=heading_row,
            required_labels=("TASK / GOAL", "CATEGORY", "STATUS", "NOTES"),
            stop_row=next_heading_row,
        )
        start_row = header_row + 1
        end_row = next_heading_row - 1
        header_map = self._header_map(
            worksheet,
            header_row,
            data_start_row=start_row,
            data_end_row=end_row,
        )
        return header_row, end_row, header_map

    def _find_label_row(self, worksheet: Worksheet, label: str) -> int:
        """Find the row containing a label fragment."""

        normalized_label = self._normalize_label(label)
        for row in worksheet.iter_rows():
            for cell in row:
                if normalized_label in self._normalize_label(cell.value):
                    return cell.row

        raise PlannerWorkbookError(
            f"Missing required planner label '{label}' in sheet '{worksheet.title}'"
        )

    def _find_header_row_after(
        self,
        worksheet: Worksheet,
        start_row: int,
        required_labels: tuple[str, ...],
        stop_row: int | None = None,
    ) -> int:
        """Find a header row containing all required labels."""

        final_row = stop_row or (worksheet.max_row + 1)
        required = {self._normalize_label(label) for label in required_labels}
        for row_number in range(start_row, final_row):
            row_labels = {
                self._normalize_label(cell.value)
                for cell in worksheet[row_number]
                if cell.value is not None
            }
            if required.issubset(row_labels):
                return row_number

        raise PlannerWorkbookError(
            f"Missing header labels {required_labels} in sheet '{worksheet.title}'"
        )

    def _find_week_heading_rows(self, worksheet: Worksheet) -> list[int]:
        """Find all WEEK section heading rows."""

        heading_rows: list[int] = []
        for row_number, row in enumerate(worksheet.iter_rows(), start=1):
            for column_number, cell in enumerate(row, start=1):
                normalized = self._normalize_label(cell.value)
                if (
                    column_number == 2
                    and normalized.startswith("WEEK ")
                    and row_number not in heading_rows
                ):
                    heading_rows.append(row_number)
                    break
        return heading_rows

    def _find_next_week_heading_row(
        self,
        worksheet: Worksheet,
        start_row: int,
    ) -> int | None:
        """Find the first WEEK heading row after a given row."""

        for heading_row in self._find_week_heading_rows(worksheet):
            if heading_row >= start_row:
                return heading_row
        return None

    def _required_column(
        self,
        header_map: dict[str, int],
        labels: tuple[str, ...],
    ) -> int:
        """Find a required header column from one or more accepted labels."""

        for label in labels:
            normalized = self._normalize_label(label)
            if normalized in header_map:
                return header_map[normalized]

        raise PlannerWorkbookError(
            f"Missing required planner column: {' or '.join(labels)}"
        )

    def _read_string(
        self,
        worksheet: Worksheet,
        row_number: int,
        header_map: dict[str, int],
        label: str,
    ) -> str | None:
        """Read a string value from a row by normalized header label."""

        column = header_map.get(self._normalize_label(label))
        if column is None:
            return None
        return self._string_or_none(worksheet.cell(row_number, column).value)

    def _cell_references(
        self,
        worksheet: Worksheet,
        row_number: int,
        header_map: dict[str, int],
        fields: dict[str, tuple[str, ...]],
    ) -> dict[str, str]:
        """Build source cell references for parsed fields."""

        references: dict[str, str] = {}
        for field_name, labels in fields.items():
            for label in labels:
                column = header_map.get(self._normalize_label(label))
                if column is not None:
                    references[field_name] = (
                        f"{get_column_letter(column)}{row_number}"
                    )
                    break
        return references

    def _looks_like_month_sheet(self, sheet_name: str) -> bool:
        """Return whether a sheet name looks like a planner month."""

        parts = sheet_name.split()
        return len(parts) == 2 and parts[0].isalpha() and parts[1].isdigit()

    def _normalize_lookup(self, value: str) -> str:
        """Normalize user lookup text."""

        return " ".join(value.strip().casefold().split())

    def _string_or_none(self, value: Any) -> str | None:
        """Convert non-empty cell values to stripped strings."""

        if value is None:
            return None
        text = str(value).strip()
        return text or None

    def _week_title_range(self, title: str, month: str) -> tuple[date, date]:
        normalized = " ".join(title.replace("–", "-").replace("—", "-").split())
        match = re.search(
            r"(\d{1,2})\s+([A-Za-z]{3})(?:\s+(\d{4}))?\s*-\s*"
            r"(\d{1,2})\s+([A-Za-z]{3})\s+(\d{4})",
            normalized,
        )
        if match is None:
            raise PlannerWorkbookError(f"Cannot parse week dates: {title}")
        end_year = int(match.group(6))
        if match.group(3):
            start_year = int(match.group(3))
        else:
            start_month = datetime.strptime(match.group(2), "%b").month
            end_month = datetime.strptime(match.group(5), "%b").month
            start_year = end_year - 1 if start_month > end_month else end_year
        start = datetime.strptime(
            f"{match.group(1)} {match.group(2)} {start_year}",
            "%d %b %Y",
        ).date()
        end = datetime.strptime(
            f"{match.group(4)} {match.group(5)} {match.group(6)}",
            "%d %b %Y",
        ).date()
        return start, end

    def _parse_dated_block(
        self,
        value: str,
    ) -> tuple[int, str | None, str | None, str | None, bool]:
        fixed = re.fullmatch(
            r"\s*(\d{1,2}:\d{2})\s*[-–]\s*(\d{1,2}:\d{2})\s*",
            value,
        )
        if fixed:
            start = datetime.strptime(fixed.group(1), "%H:%M")
            end = datetime.strptime(fixed.group(2), "%H:%M")
            minutes = max(1, int((end - start).total_seconds() // 60))
            return minutes, None, fixed.group(1), fixed.group(2), True
        flexible = re.search(
            r"(\d+)\s*min(?:ute)?s?\s*(?:[·|-]\s*)?(morning|afternoon|evening|night)?",
            value,
            re.I,
        )
        if flexible:
            return int(flexible.group(1)), flexible.group(2).casefold() if flexible.group(2) else None, None, None, False
        return 60, None, None, None, False
