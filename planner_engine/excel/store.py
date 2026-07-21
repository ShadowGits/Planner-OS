"""Low-level workbook orchestration for Excel Planner Store."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from shutil import copy2
from typing import Any

from openpyxl import load_workbook as openpyxl_load_workbook
from openpyxl.workbook.workbook import Workbook

from planner_engine.excel.layout import LayoutMixin, SheetLayout
from planner_engine.excel.reader import ReaderMixin
from planner_engine.excel.writer import WriterMixin


class ExcelPlannerStore(WriterMixin, ReaderMixin, LayoutMixin):
    """Read, write, save, and back up an Excel planner workbook."""

    def __init__(self, planner_path: str | Path, backup_dir: str | Path) -> None:
        self.planner_path = Path(planner_path)
        self.backup_dir = Path(backup_dir)
        self._sheet_layout_cache: dict[str, SheetLayout] = {}
        self._read_cache: Workbook | None = None
        self._read_cache_key: tuple[int, int] | None = None

    def load_workbook(self) -> Workbook:
        """Load a fresh, mutable copy of the planner workbook for writing.

        Writes always get their own workbook, never the shared read cache, so a
        mutation in progress can never corrupt cached read state.
        """

        if not self.planner_path.exists():
            raise FileNotFoundError(f"Planner not found: {self.planner_path}")

        return openpyxl_load_workbook(self.planner_path, data_only=False)

    def load_workbook_read_only(self) -> Workbook:
        """Return a parsed workbook for reads, parsing the file at most once.

        A single tool call reads the workbook many times (months, dated tasks,
        month plans...), and each read used to re-open and re-parse the whole
        .xlsx from disk — the dominant source of latency. We parse once and
        reuse the result until the file actually changes on disk or a write
        invalidates the cache.

        A normal (non-streaming) workbook is cached on purpose: repeated random
        cell reads stay cheap, and openpyxl's Workbook.close() is a no-op for
        normal workbooks, so the readers' existing close() calls leave the cache
        intact. Reads never mutate the workbook, so sharing one instance across
        reads is safe.
        """

        if not self.planner_path.exists():
            raise FileNotFoundError(f"Planner not found: {self.planner_path}")

        stat = self.planner_path.stat()
        key = (stat.st_mtime_ns, stat.st_size)
        if self._read_cache is not None and self._read_cache_key == key:
            return self._read_cache

        workbook = openpyxl_load_workbook(self.planner_path, data_only=False)
        self._read_cache = workbook
        self._read_cache_key = key
        return workbook

    def _invalidate_read_cache(self) -> None:
        """Drop the cached read workbook after the file changes underneath it."""

        self._read_cache = None
        self._read_cache_key = None

    def save_workbook(self, workbook: Workbook) -> None:
        """Save a workbook back to the planner path."""

        workbook.save(self.planner_path)
        self._invalidate_read_cache()

    def create_backup(self) -> Path:
        """Create a collision-safe timestamped copy of the planner workbook."""

        if not self.planner_path.exists():
            raise FileNotFoundError(f"Planner not found: {self.planner_path}")

        self.backup_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        backup_path = self.backup_dir / f"{self.planner_path.stem}_{timestamp}.xlsx"
        suffix = 1
        while backup_path.exists():
            backup_path = (
                self.backup_dir
                / f"{self.planner_path.stem}_{timestamp}_{suffix}.xlsx"
            )
            suffix += 1
        copy2(self.planner_path, backup_path)
        return backup_path

    def restore_backup(self, backup_path: str | Path) -> None:
        """Restore the planner workbook from a backup copy."""

        copy2(Path(backup_path), self.planner_path)
        self._invalidate_read_cache()

    def read_cell(self, sheet: str, cell: str) -> Any:
        """Read a value from a worksheet cell."""

        workbook = self.load_workbook()
        try:
            worksheet = workbook[sheet]
            return worksheet[cell].value
        finally:
            workbook.close()

    def write_cell(self, sheet: str, cell: str, value: Any) -> None:
        """Write a value to a worksheet cell and save the workbook."""

        workbook = self.load_workbook()
        try:
            worksheet = workbook[sheet]
            worksheet[cell].value = value
            self.save_workbook(workbook)
        finally:
            workbook.close()

    def write_cells(self, updates: list[tuple[str, str, Any]]) -> None:
        """Write multiple worksheet cells in one workbook save."""

        workbook = self.load_workbook()
        try:
            for sheet, cell, value in updates:
                workbook[sheet][cell].value = value
            self.save_workbook(workbook)
        finally:
            workbook.close()
