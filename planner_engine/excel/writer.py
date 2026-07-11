"""Semantic write support for Excel planner workbooks."""

from __future__ import annotations

from copy import copy
from datetime import date, datetime
import re
from typing import Any

from openpyxl.formula import Tokenizer
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet

from planner_engine.excel.layout import PlannerWorkbookError


class WriterMixin:
    """Append planner concepts to workbook sections."""

    def append_monthly_goals(
        self,
        month: str,
        goals: list[dict[str, Any]],
    ) -> int:
        """Back up the workbook, then append monthly goal rows."""

        if not goals:
            return 0

        self.create_backup()
        return self._append_monthly_goals_without_backup(month, goals)

    def _append_monthly_goals_without_backup(
        self,
        month: str,
        goals: list[dict[str, Any]],
    ) -> int:
        """Append monthly goal rows after backup safety has been handled."""

        workbook = self.load_workbook()
        try:
            worksheet = self._get_month_worksheet(workbook, month)
            header_row, end_row, header_map = self._monthly_goal_table(worksheet)
            task_column = self._required_column(
                header_map,
                ("GOAL / TASK", "TASK / GOAL"),
            )
            rows = self._blank_rows(
                worksheet=worksheet,
                task_column=task_column,
                start_row=header_row + 1,
                end_row=end_row,
                count=len(goals),
            )
            for row_number, goal in zip(rows, goals):
                self._write_row_values(
                    worksheet,
                    row_number,
                    header_map,
                    {
                        "GOAL / TASK": goal.get("goal"),
                        "CATEGORY": goal.get("category"),
                        "PRIORITY": goal.get("priority"),
                        "TARGET WEEK": goal.get("target_week"),
                        "STATUS": goal.get("status"),
                        "NOTES": goal.get("notes"),
                    },
                )
            self.save_workbook(workbook)
            return len(goals)
        finally:
            workbook.close()

    def append_weekly_tasks(
        self,
        month: str,
        week_tasks: list[dict[str, Any]],
    ) -> int:
        """Back up the workbook, then append weekly task rows."""

        if not week_tasks:
            return 0

        self.create_backup()
        return self._append_weekly_tasks_without_backup(month, week_tasks)

    def _append_weekly_tasks_without_backup(
        self,
        month: str,
        week_tasks: list[dict[str, Any]],
    ) -> int:
        """Append weekly task rows after backup safety has been handled."""

        workbook = self.load_workbook()
        try:
            worksheet = self._get_month_worksheet(workbook, month)
            written = 0
            for week_task in week_tasks:
                week_number = int(week_task["week"])
                header_row, end_row, header_map = self._week_task_table(
                    worksheet,
                    week_number,
                )
                task_column = self._required_column(
                    header_map,
                    ("TASK / GOAL", "GOAL / TASK"),
                )
                row_number = self._weekly_task_row(
                    worksheet=worksheet,
                    task_column=task_column,
                    start_row=header_row + 1,
                    end_row=end_row,
                )
                self._write_row_values(
                    worksheet,
                    row_number,
                    header_map,
                    {
                        "TASK / GOAL": week_task.get("task"),
                        "CATEGORY": week_task.get("category"),
                        "STATUS": week_task.get("status"),
                        "NOTES": week_task.get("notes"),
                    },
                )
                written += 1
            self.save_workbook(workbook)
            return written
        finally:
            workbook.close()

    def _append_dated_tasks_without_backup(
        self,
        month: str,
        dated_tasks: list[dict[str, Any]],
    ) -> int:
        """Append exact-date tasks into weekly rows and their day columns."""

        workbook = self.load_workbook()
        try:
            worksheet = self._get_month_worksheet(workbook, month)
            written = 0
            for task in dated_tasks:
                task_date = task["date"]
                if not isinstance(task_date, date):
                    task_date = date.fromisoformat(str(task_date))
                week_number, date_column = self._dated_task_location(
                    worksheet,
                    task_date,
                )
                header_row, end_row, header_map = self._week_task_table(
                    worksheet,
                    week_number,
                )
                task_column = self._required_column(
                    header_map,
                    ("TASK / GOAL", "GOAL / TASK"),
                )
                row_number = self._weekly_task_row(
                    worksheet,
                    task_column,
                    header_row + 1,
                    end_row,
                )
                self._write_row_values(
                    worksheet,
                    row_number,
                    header_map,
                    {
                        "TASK / GOAL": task.get("title"),
                        "CATEGORY": task.get("category"),
                        "STATUS": task.get("status", "Not Started"),
                        "NOTES": task.get("notes"),
                    },
                )
                worksheet.cell(row_number, date_column).value = task["time_value"]
                written += 1
            self.save_workbook(workbook)
            return written
        finally:
            workbook.close()

    def _dated_task_location(
        self,
        worksheet: Worksheet,
        task_date: date,
    ) -> tuple[int, int]:
        """Return workbook week number and exact day column for a date."""

        for week_number, heading_row in enumerate(
            self._find_week_heading_rows(worksheet),
            start=1,
        ):
            header_row, _, _ = self._week_task_table(worksheet, week_number)
            for column in range(4, 11):
                label = self._normalize_label(worksheet.cell(header_row, column).value)
                if task_date.strftime("%d %b").upper() in label:
                    return week_number, column
        raise PlannerWorkbookError(
            f"Date {task_date.isoformat()} is not represented in sheet '{worksheet.title}'"
        )

    def _weekly_task_row(
        self,
        worksheet: Worksheet,
        task_column: int,
        start_row: int,
        end_row: int,
    ) -> int:
        """Return an empty weekly row, expanding the section when necessary."""

        try:
            return self._blank_rows(
                worksheet=worksheet,
                task_column=task_column,
                start_row=start_row,
                end_row=end_row,
                count=1,
            )[0]
        except PlannerWorkbookError:
            insertion_row = end_row + 1
            self._insert_weekly_task_row(worksheet, insertion_row)
            return insertion_row

    def _insert_weekly_task_row(
        self,
        worksheet: Worksheet,
        insertion_row: int,
    ) -> None:
        """Insert a formatted task row before the next week heading."""

        template_row = insertion_row - 1
        source_merges = [
            copy(merged_range)
            for merged_range in worksheet.merged_cells.ranges
            if merged_range.min_row == template_row
            and merged_range.max_row == template_row
        ]
        affected_merges = [
            copy(merged_range)
            for merged_range in worksheet.merged_cells.ranges
            if merged_range.max_row >= insertion_row
        ]
        for merged_range in affected_merges:
            worksheet.unmerge_cells(str(merged_range))

        self._shift_formula_references_for_insert(worksheet, insertion_row)
        self._shift_data_validations_for_insert(
            worksheet,
            insertion_row=insertion_row,
            template_row=template_row,
        )
        worksheet.insert_rows(insertion_row, 1)
        self._copy_task_row(worksheet, template_row, insertion_row)

        for merged_range in affected_merges:
            min_row = merged_range.min_row
            max_row = merged_range.max_row
            if min_row >= insertion_row:
                min_row += 1
                max_row += 1
            else:
                max_row += 1
            worksheet.merge_cells(
                start_row=min_row,
                start_column=merged_range.min_col,
                end_row=max_row,
                end_column=merged_range.max_col,
            )

        for merged_range in source_merges:
            worksheet.merge_cells(
                start_row=insertion_row,
                start_column=merged_range.min_col,
                end_row=insertion_row,
                end_column=merged_range.max_col,
            )

    def _copy_task_row(
        self,
        worksheet: Worksheet,
        source_row: int,
        target_row: int,
    ) -> None:
        """Copy row presentation and template formulas without task values."""

        source_dimension = worksheet.row_dimensions[source_row]
        target_dimension = worksheet.row_dimensions[target_row]
        target_dimension.height = source_dimension.height
        target_dimension.hidden = source_dimension.hidden
        target_dimension.outlineLevel = source_dimension.outlineLevel

        for column in range(1, worksheet.max_column + 1):
            source_cell = worksheet.cell(source_row, column)
            target_cell = worksheet.cell(target_row, column)
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
            if isinstance(source_cell.value, str) and source_cell.value.startswith("="):
                try:
                    target_cell.value = Translator(
                        source_cell.value,
                        origin=source_cell.coordinate,
                    ).translate_formula(target_cell.coordinate)
                except ValueError:
                    target_cell.value = source_cell.value

    def _shift_formula_references_for_insert(
        self,
        worksheet: Worksheet,
        insertion_row: int,
    ) -> None:
        """Shift local A1 references as an Excel row insertion would."""

        for row in worksheet.iter_rows():
            for cell in row:
                if not (
                    isinstance(cell.value, str)
                    and cell.value.startswith("=")
                ):
                    continue
                tokenizer = Tokenizer(cell.value)
                for token in tokenizer.items:
                    if token.type == "OPERAND" and token.subtype == "RANGE":
                        token.value = self._shift_formula_range_token(
                            token.value,
                            worksheet.title,
                            insertion_row,
                        )
                cell.value = "=" + "".join(token.value for token in tokenizer.items)

    def _shift_formula_range_token(
        self,
        token: str,
        sheet_name: str,
        insertion_row: int,
    ) -> str:
        """Shift one local cell or cell-range formula token."""

        match = re.fullmatch(
            r"(?:(?P<sheet>'(?:[^']|'')+'|[^!]+)!)?"
            r"(?P<first>\$?[A-Z]{1,3}\$?\d+)"
            r"(?::(?P<second>\$?[A-Z]{1,3}\$?\d+))?",
            token,
        )
        if match is None:
            return token
        referenced_sheet = match.group("sheet")
        if referenced_sheet:
            normalized_sheet = referenced_sheet.strip("'").replace("''", "'")
            if normalized_sheet.casefold() != sheet_name.casefold():
                return token

        first = match.group("first")
        second = match.group("second")
        first_row = self._formula_reference_row(first)
        if second is None:
            shifted_first = (
                self._replace_formula_reference_row(first, first_row + 1)
                if first_row >= insertion_row
                else first
            )
            return self._formula_token_prefix(referenced_sheet) + shifted_first

        second_row = self._formula_reference_row(second)
        if insertion_row <= first_row:
            first = self._replace_formula_reference_row(first, first_row + 1)
            second = self._replace_formula_reference_row(second, second_row + 1)
        elif first_row < insertion_row <= second_row:
            second = self._replace_formula_reference_row(second, second_row + 1)
        return self._formula_token_prefix(referenced_sheet) + f"{first}:{second}"

    def _formula_reference_row(self, reference: str) -> int:
        match = re.fullmatch(r"\$?[A-Z]{1,3}\$?(?P<row>\d+)", reference)
        if match is None:
            raise ValueError(f"Unsupported formula reference: {reference}")
        return int(match.group("row"))

    def _replace_formula_reference_row(self, reference: str, row: int) -> str:
        return re.sub(r"\d+$", str(row), reference)

    def _formula_token_prefix(self, referenced_sheet: str | None) -> str:
        return f"{referenced_sheet}!" if referenced_sheet else ""

    def _shift_data_validations_for_insert(
        self,
        worksheet: Worksheet,
        insertion_row: int,
        template_row: int,
    ) -> None:
        """Shift later validation ranges and copy template-row validation."""

        for validation in worksheet.data_validations.dataValidation:
            adjusted_ranges: list[CellRange] = []
            template_columns: list[tuple[int, int]] = []
            for cell_range in list(validation.ranges.ranges):
                if cell_range.min_row <= template_row <= cell_range.max_row:
                    template_columns.append(
                        (cell_range.min_col, cell_range.max_col)
                    )

                min_row = cell_range.min_row
                max_row = cell_range.max_row
                if min_row >= insertion_row:
                    min_row += 1
                    max_row += 1
                elif min_row < insertion_row <= max_row:
                    max_row += 1
                adjusted_ranges.append(
                    CellRange(
                        min_col=cell_range.min_col,
                        min_row=min_row,
                        max_col=cell_range.max_col,
                        max_row=max_row,
                    )
                )

            validation.ranges = []
            for cell_range in adjusted_ranges:
                validation.add(str(cell_range))
            for min_col, max_col in template_columns:
                validation.add(
                    f"{get_column_letter(min_col)}{insertion_row}:"
                    f"{get_column_letter(max_col)}{insertion_row}"
                )

    def _blank_rows(
        self,
        worksheet: Worksheet,
        task_column: int,
        start_row: int,
        end_row: int,
        count: int,
    ) -> list[int]:
        """Find blank data rows in a planner section."""

        rows: list[int] = []
        for row_number in range(start_row, end_row + 1):
            value = self._string_or_none(worksheet.cell(row_number, task_column).value)
            if value and value.startswith("↳"):
                continue
            if value:
                continue
            rows.append(row_number)
            if len(rows) == count:
                return rows

        raise PlannerWorkbookError(
            f"Not enough blank rows in sheet '{worksheet.title}'"
        )

    def _write_row_values(
        self,
        worksheet: Worksheet,
        row_number: int,
        header_map: dict[str, int],
        values: dict[str, Any],
    ) -> None:
        """Write values into a planner row by semantic heading."""

        for heading, value in values.items():
            column = header_map.get(self._normalize_label(heading))
            if column is not None:
                worksheet.cell(row_number, column).value = value
