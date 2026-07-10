"""Excel workbook storage for Planner Engine."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from pathlib import PurePosixPath
from shutil import copy2
from typing import Any
from zipfile import ZipFile
import re
import xml.etree.ElementTree as ElementTree

from openpyxl import load_workbook as openpyxl_load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from planner_engine.models import (
    MonthPlan,
    MonthlyGoal,
    PlannerTask,
    Priority,
    TaskStatus,
    WeekSection,
)


class PlannerWorkbookError(ValueError):
    """Raised when planner workbook structure cannot be parsed."""


@dataclass(frozen=True)
class CellRange:
    """A cell range parsed from workbook XML."""

    min_col: int
    min_row: int
    max_col: int
    max_row: int
    reference: str

    def contains(self, row: int, column: int) -> bool:
        """Return whether a cell is inside this range."""

        return (
            self.min_row <= row <= self.max_row
            and self.min_col <= column <= self.max_col
        )

    def intersects_rows(self, start_row: int, end_row: int) -> bool:
        """Return whether the range overlaps a row interval."""

        return not (self.max_row < start_row or self.min_row > end_row)

    def intersects_columns(self, start_column: int, end_column: int) -> bool:
        """Return whether the range overlaps a column interval."""

        return not (self.max_col < start_column or self.min_col > end_column)


@dataclass(frozen=True)
class SheetLayout:
    """Merged-cell and validation layout for a sheet."""

    merged_ranges: tuple[CellRange, ...]
    data_validation_ranges: tuple[CellRange, ...]


class ExcelPlannerStore:
    """Read, write, save, and back up an Excel planner workbook."""

    def __init__(self, planner_path: str | Path, backup_dir: str | Path) -> None:
        self.planner_path = Path(planner_path)
        self.backup_dir = Path(backup_dir)
        self._sheet_layout_cache: dict[str, SheetLayout] = {}

    def load_workbook(self) -> Workbook:
        """Load the planner workbook with formatting preserved."""

        if not self.planner_path.exists():
            raise FileNotFoundError(f"Planner not found: {self.planner_path}")

        return openpyxl_load_workbook(self.planner_path, data_only=False)

    def load_workbook_read_only(self) -> Workbook:
        """Load the planner workbook for read-only parsing."""

        if not self.planner_path.exists():
            raise FileNotFoundError(f"Planner not found: {self.planner_path}")

        return openpyxl_load_workbook(
            self.planner_path,
            read_only=True,
            data_only=False,
        )

    def save_workbook(self, workbook: Workbook) -> None:
        """Save a workbook back to the planner path."""

        workbook.save(self.planner_path)

    def create_backup(self) -> Path:
        """Create a timestamped copy of the planner workbook."""

        if not self.planner_path.exists():
            raise FileNotFoundError(f"Planner not found: {self.planner_path}")

        self.backup_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = self.backup_dir / f"{self.planner_path.stem}_{timestamp}.xlsx"
        copy2(self.planner_path, backup_path)
        return backup_path

    def read_cell(self, sheet: str, cell: str) -> Any:
        """Read a value from a worksheet cell."""

        workbook = self.load_workbook()
        try:
            worksheet = workbook[sheet]
            return worksheet[cell].value
        finally:
            workbook.close()

    def write_cell(self, sheet: str, cell: str, value: Any) -> None:
        """Write a value to a worksheet cell and save the workbook."""

        workbook = self.load_workbook()
        try:
            worksheet = workbook[sheet]
            worksheet[cell].value = value
            self.save_workbook(workbook)
        finally:
            workbook.close()

    def list_months(self) -> list[str]:
        """List monthly planner sheet names."""

        workbook = self.load_workbook_read_only()
        try:
            return [
                sheet_name
                for sheet_name in workbook.sheetnames
                if self._looks_like_month_sheet(sheet_name)
            ]
        finally:
            workbook.close()

    def read_month_plan(self, month: str) -> MonthPlan:
        """Read a full parsed month plan without modifying the workbook."""

        workbook = self.load_workbook_read_only()
        try:
            worksheet = self._get_month_worksheet(workbook, month)
            return MonthPlan(
                month=worksheet.title,
                sheet_name=worksheet.title,
                monthly_goals=self._parse_monthly_goals(worksheet),
                week_sections=self._parse_week_sections(worksheet),
            )
        finally:
            workbook.close()

    def read_monthly_goals(self, month: str) -> list[MonthlyGoal]:
        """Read monthly goals from a month sheet."""

        return self.read_month_plan(month).monthly_goals

    def read_week_sections(self, month: str) -> list[WeekSection]:
        """Read weekly sections from a month sheet."""

        return self.read_month_plan(month).week_sections

    def find_task(self, month: str, task_name: str) -> PlannerTask | MonthlyGoal | None:
        """Find a task or goal by case-insensitive name."""

        normalized_task_name = self._normalize_lookup(task_name)
        month_plan = self.read_month_plan(month)

        for goal in month_plan.monthly_goals:
            if self._normalize_lookup(goal.name) == normalized_task_name:
                return goal

        for week_section in month_plan.week_sections:
            for task in week_section.tasks:
                if self._normalize_lookup(task.name) == normalized_task_name:
                    return task

        return None

    def _get_month_worksheet(self, workbook: Workbook, month: str) -> Worksheet:
        """Return a month worksheet, or raise a clear parsing error."""

        requested_month = self._normalize_lookup(month)
        for sheet_name in workbook.sheetnames:
            if self._normalize_lookup(sheet_name) == requested_month:
                return workbook[sheet_name]

        raise PlannerWorkbookError(f"Unknown planner month: {month}")

    def _parse_monthly_goals(self, worksheet: Worksheet) -> list[MonthlyGoal]:
        """Parse monthly goals using heading labels."""

        section_row = self._find_label_row(worksheet, "MONTHLY GOALS")
        header_row = self._find_header_row_after(
            worksheet,
            start_row=section_row,
            required_labels=("GOAL / TASK", "CATEGORY", "PRIORITY", "TARGET WEEK"),
        )
        next_week_row = self._find_next_week_heading_row(
            worksheet,
            start_row=header_row + 1,
        )
        end_row = (next_week_row - 1) if next_week_row else worksheet.max_row
        header_map = self._header_map(
            worksheet,
            header_row,
            data_start_row=header_row + 1,
            data_end_row=end_row,
        )
        task_column = self._required_column(header_map, ("GOAL / TASK", "TASK / GOAL"))

        goals: list[MonthlyGoal] = []
        for row_number in range(header_row + 1, end_row + 1):
            name = self._string_or_none(worksheet.cell(row_number, task_column).value)
            if not name:
                continue

            category = self._read_string(worksheet, row_number, header_map, "CATEGORY")
            raw_priority = self._read_string(
                worksheet,
                row_number,
                header_map,
                "PRIORITY",
            )
            target_week = self._read_string(
                worksheet,
                row_number,
                header_map,
                "TARGET WEEK",
            )
            raw_status = self._read_string(worksheet, row_number, header_map, "STATUS")
            notes = self._read_string(worksheet, row_number, header_map, "NOTES")

            goals.append(
                MonthlyGoal(
                    name=name,
                    category=category,
                    priority=Priority.from_value(raw_priority),
                    target_week=target_week,
                    status=TaskStatus.from_value(raw_status),
                    notes=notes,
                    sheet_name=worksheet.title,
                    row_number=row_number,
                    cell_references=self._cell_references(
                        worksheet,
                        row_number,
                        header_map,
                        {
                            "name": ("GOAL / TASK", "TASK / GOAL"),
                            "category": ("CATEGORY",),
                            "priority": ("PRIORITY",),
                            "target_week": ("TARGET WEEK",),
                            "status": ("STATUS",),
                            "notes": ("NOTES",),
                        },
                    ),
                    raw_priority=raw_priority,
                    raw_status=raw_status,
                )
            )

        return goals

    def _parse_week_sections(self, worksheet: Worksheet) -> list[WeekSection]:
        """Parse all weekly sections using WEEK labels."""

        week_heading_rows = self._find_week_heading_rows(worksheet)
        week_sections: list[WeekSection] = []

        for index, heading_row in enumerate(week_heading_rows):
            next_heading_row = (
                week_heading_rows[index + 1]
                if index + 1 < len(week_heading_rows)
                else worksheet.max_row + 1
            )
            header_row = self._find_header_row_after(
                worksheet,
                start_row=heading_row,
                required_labels=("TASK / GOAL", "CATEGORY", "STATUS", "NOTES"),
                stop_row=next_heading_row,
            )
            title = self._string_or_none(worksheet.cell(heading_row, 2).value) or ""
            week_name = title.split("·", maxsplit=1)[0].strip()
            start_row = header_row + 1
            end_row = next_heading_row - 1
            header_map = self._header_map(
                worksheet,
                header_row,
                data_start_row=start_row,
                data_end_row=end_row,
            )

            tasks = self._parse_week_tasks(
                worksheet=worksheet,
                week_name=week_name,
                header_row=header_row,
                start_row=start_row,
                end_row=end_row,
                header_map=header_map,
            )
            week_sections.append(
                WeekSection(
                    name=week_name,
                    title=title,
                    sheet_name=worksheet.title,
                    heading_row=heading_row,
                    header_row=header_row,
                    start_row=start_row,
                    end_row=end_row,
                    tasks=tasks,
                    cell_references={
                        "heading": worksheet.cell(heading_row, 2).coordinate,
                    },
                )
            )

        return week_sections

    def _parse_week_tasks(
        self,
        worksheet: Worksheet,
        week_name: str,
        header_row: int,
        start_row: int,
        end_row: int,
        header_map: dict[str, int],
    ) -> list[PlannerTask]:
        """Parse task rows in a weekly section."""

        task_column = self._required_column(header_map, ("TASK / GOAL", "GOAL / TASK"))
        tasks: list[PlannerTask] = []

        for row_number in range(start_row, end_row + 1):
            name = self._string_or_none(worksheet.cell(row_number, task_column).value)
            if not name or name.startswith("↳"):
                continue

            category = self._read_string(worksheet, row_number, header_map, "CATEGORY")
            raw_status = self._read_string(worksheet, row_number, header_map, "STATUS")
            notes = self._read_string(worksheet, row_number, header_map, "NOTES")

            tasks.append(
                PlannerTask(
                    name=name,
                    category=category,
                    priority=Priority.UNKNOWN,
                    status=TaskStatus.from_value(raw_status),
                    notes=notes,
                    sheet_name=worksheet.title,
                    row_number=row_number,
                    week_name=week_name,
                    cell_references=self._cell_references(
                        worksheet,
                        row_number,
                        header_map,
                        {
                            "name": ("TASK / GOAL", "GOAL / TASK"),
                            "category": ("CATEGORY",),
                            "status": ("STATUS",),
                            "notes": ("NOTES",),
                        },
                    ),
                    raw_status=raw_status,
                )
            )

        return tasks

    def _find_label_row(self, worksheet: Worksheet, label: str) -> int:
        """Find the row containing a label fragment."""

        normalized_label = self._normalize_label(label)
        for row in worksheet.iter_rows():
            for cell in row:
                if normalized_label in self._normalize_label(cell.value):
                    return cell.row

        raise PlannerWorkbookError(
            f"Missing required planner label '{label}' in sheet '{worksheet.title}'"
        )

    def _find_header_row_after(
        self,
        worksheet: Worksheet,
        start_row: int,
        required_labels: tuple[str, ...],
        stop_row: int | None = None,
    ) -> int:
        """Find a header row containing all required labels."""

        final_row = stop_row or (worksheet.max_row + 1)
        required = {self._normalize_label(label) for label in required_labels}
        for row_number in range(start_row, final_row):
            row_labels = {
                self._normalize_label(cell.value)
                for cell in worksheet[row_number]
                if cell.value is not None
            }
            if required.issubset(row_labels):
                return row_number

        raise PlannerWorkbookError(
            f"Missing header labels {required_labels} in sheet '{worksheet.title}'"
        )

    def _find_week_heading_rows(self, worksheet: Worksheet) -> list[int]:
        """Find all WEEK section heading rows."""

        heading_rows: list[int] = []
        for row_number, row in enumerate(worksheet.iter_rows(), start=1):
            for column_number, cell in enumerate(row, start=1):
                normalized = self._normalize_label(cell.value)
                if (
                    column_number == 2
                    and normalized.startswith("WEEK ")
                    and row_number not in heading_rows
                ):
                    heading_rows.append(row_number)
                    break
        return heading_rows

    def _find_next_week_heading_row(
        self,
        worksheet: Worksheet,
        start_row: int,
    ) -> int | None:
        """Find the first WEEK heading row after a given row."""

        for heading_row in self._find_week_heading_rows(worksheet):
            if heading_row >= start_row:
                return heading_row
        return None

    def _header_map(
        self,
        worksheet: Worksheet,
        header_row: int,
        data_start_row: int,
        data_end_row: int,
    ) -> dict[str, int]:
        """Map normalized header labels to actual entry column indexes."""

        headers: dict[str, int] = {}
        for cell in worksheet[header_row]:
            normalized = self._normalize_label(cell.value)
            if normalized:
                headers[normalized] = self._entry_column_for_header(
                    sheet_name=worksheet.title,
                    header_row=header_row,
                    header_column=cell.column,
                    data_start_row=data_start_row,
                    data_end_row=data_end_row,
                )
        return headers

    def _entry_column_for_header(
        self,
        sheet_name: str,
        header_row: int,
        header_column: int,
        data_start_row: int,
        data_end_row: int,
    ) -> int:
        """Resolve a visible header cell to the workbook's data-entry column."""

        layout = self._sheet_layout(sheet_name)
        header_range = self._merged_range_containing(
            layout,
            row=header_row,
            column=header_column,
        )
        if header_range is None:
            return header_column

        for validation_range in layout.data_validation_ranges:
            if (
                validation_range.intersects_rows(data_start_row, data_end_row)
                and validation_range.intersects_columns(
                    header_range.min_col,
                    header_range.max_col,
                )
            ):
                return max(validation_range.min_col, header_range.min_col)

        for merged_range in layout.merged_ranges:
            if (
                merged_range.min_col == header_range.min_col
                and merged_range.max_col == header_range.max_col
                and merged_range.intersects_rows(data_start_row, data_end_row)
            ):
                return merged_range.min_col

        return header_column

    def _merged_range_containing(
        self,
        layout: SheetLayout,
        row: int,
        column: int,
    ) -> CellRange | None:
        """Return the merged range containing a cell, if any."""

        for merged_range in layout.merged_ranges:
            if merged_range.contains(row, column):
                return merged_range
        return None

    def _sheet_layout(self, sheet_name: str) -> SheetLayout:
        """Load sheet merged ranges and validation ranges from workbook XML."""

        if sheet_name not in self._sheet_layout_cache:
            self._sheet_layout_cache[sheet_name] = self._load_sheet_layout(sheet_name)
        return self._sheet_layout_cache[sheet_name]

    def _load_sheet_layout(self, sheet_name: str) -> SheetLayout:
        """Read workbook XML layout metadata without modifying the workbook."""

        namespace = {
            "m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        }
        rel_namespace = {
            "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
        }

        with ZipFile(self.planner_path) as workbook_archive:
            workbook_root = ElementTree.fromstring(
                workbook_archive.read("xl/workbook.xml")
            )
            relationship_root = ElementTree.fromstring(
                workbook_archive.read("xl/_rels/workbook.xml.rels")
            )
            relationships = {
                relationship.attrib["Id"]: self._resolve_workbook_target(
                    relationship.attrib["Target"]
                )
                for relationship in relationship_root.findall(
                    "pr:Relationship",
                    rel_namespace,
                )
            }

            sheet_target: str | None = None
            relationship_key = (
                "{http://schemas.openxmlformats.org/officeDocument/2006/"
                "relationships}id"
            )
            for sheet in workbook_root.findall(".//m:sheet", namespace):
                if sheet.attrib["name"] == sheet_name:
                    sheet_target = relationships[sheet.attrib[relationship_key]]
                    break

            if sheet_target is None:
                raise PlannerWorkbookError(f"Unknown planner month: {sheet_name}")

            sheet_root = ElementTree.fromstring(workbook_archive.read(sheet_target))
            merged_ranges = tuple(
                self._parse_cell_range(merge_cell.attrib["ref"])
                for merge_cell in sheet_root.findall(
                    ".//m:mergeCells/m:mergeCell",
                    namespace,
                )
            )
            data_validation_ranges: list[CellRange] = []
            for data_validation in sheet_root.findall(
                ".//m:dataValidations/m:dataValidation",
                namespace,
            ):
                for reference in data_validation.attrib.get("sqref", "").split():
                    data_validation_ranges.append(self._parse_cell_range(reference))

        return SheetLayout(
            merged_ranges=merged_ranges,
            data_validation_ranges=tuple(data_validation_ranges),
        )

    def _resolve_workbook_target(self, target: str) -> str:
        """Resolve workbook relationship targets to archive paths."""

        if target.startswith("/xl/"):
            return target.lstrip("/")
        if target.startswith("xl/"):
            return target
        return str(PurePosixPath("xl") / target)

    def _parse_cell_range(self, reference: str) -> CellRange:
        """Parse an Excel cell or cell range reference."""

        if ":" not in reference:
            column, row = self._split_cell_reference(reference)
            column_number = self._column_number(column)
            return CellRange(
                min_col=column_number,
                min_row=row,
                max_col=column_number,
                max_row=row,
                reference=reference,
            )

        start, end = reference.split(":", maxsplit=1)
        start_column, start_row = self._split_cell_reference(start)
        end_column, end_row = self._split_cell_reference(end)
        return CellRange(
            min_col=self._column_number(start_column),
            min_row=start_row,
            max_col=self._column_number(end_column),
            max_row=end_row,
            reference=reference,
        )

    def _split_cell_reference(self, reference: str) -> tuple[str, int]:
        """Split a cell reference into column letters and row number."""

        match = re.fullmatch(r"([A-Z]+)(\d+)", reference)
        if match is None:
            raise PlannerWorkbookError(f"Invalid cell reference: {reference}")
        return match.group(1), int(match.group(2))

    def _column_number(self, column: str) -> int:
        """Convert Excel column letters to a 1-based column number."""

        column_number = 0
        for character in column:
            column_number = column_number * 26 + ord(character) - 64
        return column_number

    def _required_column(
        self,
        header_map: dict[str, int],
        labels: tuple[str, ...],
    ) -> int:
        """Find a required header column from one or more accepted labels."""

        for label in labels:
            normalized = self._normalize_label(label)
            if normalized in header_map:
                return header_map[normalized]

        raise PlannerWorkbookError(
            f"Missing required planner column: {' or '.join(labels)}"
        )

    def _read_string(
        self,
        worksheet: Worksheet,
        row_number: int,
        header_map: dict[str, int],
        label: str,
    ) -> str | None:
        """Read a string value from a row by normalized header label."""

        column = header_map.get(self._normalize_label(label))
        if column is None:
            return None
        return self._string_or_none(worksheet.cell(row_number, column).value)

    def _cell_references(
        self,
        worksheet: Worksheet,
        row_number: int,
        header_map: dict[str, int],
        fields: dict[str, tuple[str, ...]],
    ) -> dict[str, str]:
        """Build source cell references for parsed fields."""

        references: dict[str, str] = {}
        for field_name, labels in fields.items():
            for label in labels:
                column = header_map.get(self._normalize_label(label))
                if column is not None:
                    references[field_name] = (
                        f"{get_column_letter(column)}{row_number}"
                    )
                    break
        return references

    def _looks_like_month_sheet(self, sheet_name: str) -> bool:
        """Return whether a sheet name looks like a planner month."""

        parts = sheet_name.split()
        return len(parts) == 2 and parts[0].isalpha() and parts[1].isdigit()

    def _normalize_label(self, value: Any) -> str:
        """Normalize worksheet labels and headings for matching."""

        if value is None:
            return ""
        normalized = " ".join(str(value).strip().upper().replace("\n", " ").split())
        return normalized.lstrip("▸").strip()

    def _normalize_lookup(self, value: str) -> str:
        """Normalize user lookup text."""

        return " ".join(value.strip().casefold().split())

    def _string_or_none(self, value: Any) -> str | None:
        """Convert non-empty cell values to stripped strings."""

        if value is None:
            return None
        text = str(value).strip()
        return text or None
