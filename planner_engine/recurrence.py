"""Workbook-backed recurring planner definitions and occurrence generation."""

from __future__ import annotations

import json
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook

from planner_engine.preview_contract import PreviewContract
from planner_engine.writer import Writer, WriterResult


RECURRENCE_SHEET = "_RECURRENCES"
HEADERS = ["version", "recurrence_id", "title", "frequency", "selected_weekdays", "start_date", "until_date", "count", "estimated_minutes", "preferred_daypart", "category", "status"]


@dataclass(frozen=True)
class RecurrenceRequest:
    title: str
    frequency: str
    start_date: date
    until_date: date | None = None
    count: int | None = None
    selected_weekdays: list[str] = field(default_factory=list)
    estimated_minutes: int = 30
    preferred_daypart: str | None = None
    category: str | None = None


@dataclass(frozen=True)
class RecurrencePreview:
    preview_id: str
    operation: str
    recurrence_id: str
    occurrences: list[dict]
    warnings: list[str]
    requires_confirmation: bool = True
    created_at: str = field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

    def to_dict(self) -> dict:
        return asdict(self)


class RecurrenceService:
    def __init__(self, workbook_path: str | Path, writer: Writer, preview_dir: str | Path | None = None) -> None:
        self.workbook_path = Path(workbook_path)
        self.writer = writer
        self.preview_dir = Path(preview_dir) if preview_dir is not None else None
        self._last_preview: RecurrencePreview | None = None
        self._contract = PreviewContract({"workbook": self.workbook_path})

    def preview_recurrence(self, request: RecurrenceRequest) -> RecurrencePreview:
        recurrence_id = str(uuid4())
        occurrences = [
            {
                "date": item.isoformat(),
                "title": request.title,
                "estimated_minutes": request.estimated_minutes,
                "preferred_daypart": request.preferred_daypart,
                "category": request.category,
                "notes": f"Planner OS generated recurrence_id={recurrence_id}",
            }
            for item in self._dates(request)
        ]
        preview = RecurrencePreview(str(uuid4()), "apply_recurrence", recurrence_id, occurrences, [] if occurrences else ["No occurrences generated"])
        self._last_preview = preview
        self._save_preview(preview)
        return preview

    def apply_recurrence(self, preview: RecurrencePreview | str | None = None) -> WriterResult:
        selected = self._load_preview(preview) if isinstance(preview, str) else preview or self._last_preview
        if selected is None:
            raise ValueError("No recurrence preview supplied")
        result = self.writer.add_dated_tasks(selected.occurrences)
        if result.success:
            self._store_definition(selected)
            self._mark_applied(selected.preview_id)
        return result

    def pause_recurrence(self, recurrence_id: str) -> dict:
        return self._set_status(recurrence_id, "paused")

    def resume_recurrence(self, recurrence_id: str) -> dict:
        return self._set_status(recurrence_id, "active")

    def list_recurrences(self) -> list[dict]:
        workbook = self._load()
        try:
            if RECURRENCE_SHEET not in workbook.sheetnames:
                return []
            sheet = workbook[RECURRENCE_SHEET]
            rows = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[1]:
                    continue
                rows.append(dict(zip(HEADERS, row)))
            return rows
        finally:
            workbook.close()

    def _dates(self, request: RecurrenceRequest) -> list[date]:
        if request.count is None and request.until_date is None:
            raise ValueError("Recurring definitions require count or until_date")
        frequency = request.frequency.casefold().replace(" ", "_")
        allowed = {day.casefold() for day in request.selected_weekdays}
        cursor = request.start_date
        generated: list[date] = []
        end = request.until_date or request.start_date + timedelta(days=366)
        while cursor <= end and (request.count is None or len(generated) < request.count):
            weekday = cursor.strftime("%A").casefold()
            include = (
                frequency == "daily"
                or (frequency == "weekdays" and cursor.weekday() < 5)
                or (frequency == "weekends" and cursor.weekday() >= 5)
                or (frequency == "selected_weekdays" and weekday in allowed)
                or (frequency == "weekly" and cursor.weekday() == request.start_date.weekday())
                or (frequency == "monthly" and cursor.day == request.start_date.day)
            )
            if include:
                generated.append(cursor)
            cursor += timedelta(days=1)
        return generated

    def _store_definition(self, preview: RecurrencePreview) -> None:
        workbook = self._load()
        try:
            sheet = self._sheet(workbook)
            first = preview.occurrences[0] if preview.occurrences else {}
            dates = [date.fromisoformat(item["date"]) for item in preview.occurrences]
            sheet.append([
                "1",
                preview.recurrence_id,
                first.get("title", ""),
                "generated",
                "",
                min(dates).isoformat() if dates else "",
                max(dates).isoformat() if dates else "",
                len(dates),
                first.get("estimated_minutes", ""),
                first.get("preferred_daypart", ""),
                first.get("category", ""),
                "active",
            ])
            workbook.save(self.workbook_path)
        finally:
            workbook.close()

    def _save_preview(self, preview: RecurrencePreview) -> None:
        if self.preview_dir is None:
            return
        self.preview_dir.mkdir(parents=True, exist_ok=True)
        payload = self._contract.seal(
            preview.to_dict(), kind="apply_recurrence", depends_on=("workbook",)
        )
        (self.preview_dir / f"{preview.preview_id}.json").write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")

    def _load_preview(self, preview_id: str) -> RecurrencePreview:
        if self.preview_dir is None:
            raise ValueError("Recurrence preview store is not configured")
        path = self.preview_dir / f"{preview_id}.json"
        if not path.exists():
            raise ValueError("Unknown recurrence preview")
        data = json.loads(path.read_text(encoding="utf-8"))
        if data.get("applied_at"):
            raise ValueError("Recurrence preview was already applied")
        if data["operation"] != "apply_recurrence":
            raise ValueError("Recurrence preview operation does not match")
        self._contract.validate(data, kind="apply_recurrence")
        return RecurrencePreview(**{key: data[key] for key in RecurrencePreview.__dataclass_fields__ if key in data})

    def _mark_applied(self, preview_id: str) -> None:
        if self.preview_dir is None:
            return
        path = self.preview_dir / f"{preview_id}.json"
        if not path.exists():
            return
        data = json.loads(path.read_text(encoding="utf-8"))
        data["applied_at"] = datetime.now(timezone.utc).isoformat()
        path.write_text(json.dumps(data, indent=2, sort_keys=True) + "\n", encoding="utf-8")

    def _set_status(self, recurrence_id: str, status: str) -> dict:
        workbook = self._load()
        try:
            sheet = self._sheet(workbook)
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row, 2).value == recurrence_id:
                    sheet.cell(row, 12).value = status
                    workbook.save(self.workbook_path)
                    return {"success": True, "recurrence_id": recurrence_id, "status": status}
            return {"success": False, "errors": ["Unknown recurrence_id"]}
        finally:
            workbook.close()

    def _load(self):
        return load_workbook(self.workbook_path)

    def _sheet(self, workbook):
        if RECURRENCE_SHEET not in workbook.sheetnames:
            sheet = workbook.create_sheet(RECURRENCE_SHEET)
            sheet.sheet_state = "hidden"
            sheet.append(HEADERS)
        else:
            sheet = workbook[RECURRENCE_SHEET]
        return sheet
