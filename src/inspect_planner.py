from pathlib import Path
from openpyxl import load_workbook

PLANNER_PATH = Path(
    "/Users/sparry00/Library/CloudStorage/"
    "GoogleDrive-sparsh0304@gmail.com/My Drive/"
    "Life tracking/Master_Planner_Jul26_Jun27.xlsx"
)

if not PLANNER_PATH.exists():
    raise FileNotFoundError(f"Planner not found: {PLANNER_PATH}")

workbook = load_workbook(
    PLANNER_PATH,
    read_only=True,
    data_only=False,
)

print(f"Planner: {PLANNER_PATH}")
print("\nWorksheets:")

for sheet in workbook.worksheets:
    print(
        f"- {sheet.title}: "
        f"{sheet.max_row} rows × {sheet.max_column} columns"
    )

workbook.close()
