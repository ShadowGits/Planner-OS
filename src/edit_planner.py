from pathlib import Path
from datetime import datetime
from shutil import copy2
import argparse

from openpyxl import load_workbook

PLANNER = Path("/Users/sparry00/Library/CloudStorage/GoogleDrive-sparsh0304@gmail.com/My Drive/Life tracking/Master_Planner_Jul26_Jun27.xlsx")
BACKUPS = Path("backups")
BACKUPS.mkdir(exist_ok=True)

parser = argparse.ArgumentParser()
parser.add_argument("--sheet", required=True)
parser.add_argument("--cell", required=True)
parser.add_argument("--value", required=True)

args = parser.parse_args()

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
backup = BACKUPS / f"{PLANNER.stem}_{timestamp}.xlsx"
copy2(PLANNER, backup)

wb = load_workbook(PLANNER)
ws = wb[args.sheet]

ws[args.cell] = args.value

wb.save(PLANNER)
wb.close()

print(f"Backup: {backup}")
print(f"Updated {args.sheet}!{args.cell} = {args.value}")
